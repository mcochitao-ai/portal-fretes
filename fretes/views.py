# Imports organizados
import json
import os
import openpyxl
from datetime import timedelta
from openpyxl.utils import get_column_letter

from django import forms
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.db import models
from django.contrib.auth.forms import UserCreationForm, PasswordResetForm, SetPasswordForm
from django import forms
from django.contrib.auth.models import User
from django.db.models import Count, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages

from .models import Loja, FreteRequest, Destino, Transportadora, UserProfile, CotacaoFrete

# Formulário customizado para signup
class CustomUserCreationForm(UserCreationForm):
    username = forms.CharField(
        max_length=30,
        help_text="Obrigatório. 30 caracteres ou menos. Letras, números e @/./+/-/_ apenas.",
        widget=forms.TextInput(attrs={'class': 'form-control'})
    )
    email = forms.EmailField(
        required=True,
        help_text="Obrigatório. Digite um email válido.",
        widget=forms.EmailInput(attrs={'class': 'form-control'})
    )
    
    class Meta:
        model = User
        fields = ("username", "email", "password1", "password2")
    
    def save(self, commit=True):
        user = super().save(commit=False)
        user.email = self.cleaned_data["email"]
        if commit:
            user.save()
        return user

# Views principais
def signup(request):
    """View para registro de novos usuários"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'fretes/signup.html', {'form': form})


def verificar_e_criar_todas_tabelas():
    """Verifica e cria todas as tabelas e colunas necessárias"""
    from django.db import connection
    import os
    
    # Só executa se estiver usando PostgreSQL
    if not os.environ.get('DATABASE_URL'):
        return
    
    try:
        with connection.cursor() as cursor:
            print("🔧 Verificando e criando todas as tabelas...")
            
            # 1. CRIAR TABELA AUTH_USER SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS auth_user (
                    id SERIAL PRIMARY KEY,
                    password VARCHAR(128) NOT NULL,
                    last_login TIMESTAMP WITH TIME ZONE,
                    is_superuser BOOLEAN NOT NULL DEFAULT FALSE,
                    username VARCHAR(150) NOT NULL UNIQUE,
                    first_name VARCHAR(150) NOT NULL DEFAULT '',
                    last_name VARCHAR(150) NOT NULL DEFAULT '',
                    email VARCHAR(254) NOT NULL DEFAULT '',
                    is_staff BOOLEAN NOT NULL DEFAULT FALSE,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    date_joined TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
                );
            """)
            print("✅ Tabela auth_user OK")
            
            # 2. CRIAR TABELA FRETES_TRANSPORTADORA SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_transportadora (
                    id SERIAL PRIMARY KEY,
                    nome VARCHAR(255) NOT NULL,
                    email VARCHAR(254) NOT NULL
                );
            """)
            print("✅ Tabela fretes_transportadora OK")
            
            # 3. CRIAR TABELA FRETES_LOJA SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_loja (
                    id SERIAL PRIMARY KEY,
                    nome VARCHAR(255) NOT NULL,
                    endereco TEXT NOT NULL,
                    numero VARCHAR(20) DEFAULT '',
                    municipio VARCHAR(100) DEFAULT '',
                    estado VARCHAR(2) DEFAULT '',
                    cep VARCHAR(15) DEFAULT '',
                    regional VARCHAR(100) DEFAULT '',
                    latitude DECIMAL(9,6),
                    longitude DECIMAL(9,6)
                );
            """)
            print("✅ Tabela fretes_loja OK")
            
            # 4. CRIAR TABELA FRETES_FRETEREQUEST SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_freterequest (
                    id SERIAL PRIMARY KEY,
                    usuario_id INTEGER NOT NULL REFERENCES auth_user(id) ON DELETE CASCADE,
                    data_criacao TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
                    descricao TEXT DEFAULT '',
                    transportadora_selecionada_id INTEGER REFERENCES fretes_transportadora(id) ON DELETE SET NULL,
                    origem_id INTEGER REFERENCES fretes_loja(id) ON DELETE SET NULL,
                    horario_coleta VARCHAR(50) DEFAULT '',
                    observacoes_origem TEXT DEFAULT '',
                    anexo_origem VARCHAR(100) DEFAULT '',
                    tipo_veiculo VARCHAR(10) DEFAULT '',
                    precisa_ajudante BOOLEAN DEFAULT FALSE,
                    quantidade_ajudantes INTEGER DEFAULT 0,
                    nota_fiscal_emitida BOOLEAN DEFAULT FALSE,
                    anexo_nota_fiscal VARCHAR(100) DEFAULT '',
                    quem_paga_frete VARCHAR(100) DEFAULT '',
                    status VARCHAR(30) DEFAULT 'pendente',
                    centro_custo VARCHAR(100) DEFAULT '',
                    aprovador_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                    data_aprovacao TIMESTAMP WITH TIME ZONE,
                    observacoes_aprovacao TEXT DEFAULT '',
                    justificativa_rejeicao TEXT DEFAULT '',
                    motivo_cancelamento TEXT DEFAULT '',
                    usuario_cancelamento_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                    data_cancelamento TIMESTAMP WITH TIME ZONE,
                    transportadora_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                    valor_frete DECIMAL(10,2),
                    valor_pedagio DECIMAL(10,2),
                    valor_ajudante DECIMAL(10,2),
                    valor_total DECIMAL(10,2),
                    data_cotacao TIMESTAMP WITH TIME ZONE,
                    observacoes_cotacao TEXT DEFAULT '',
                    motivo_rejeicao_transportadora TEXT DEFAULT ''
                );
            """)
            print("✅ Tabela fretes_freterequest OK")
            
            # 5. CRIAR TABELA FRETES_DESTINO SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_destino (
                    id SERIAL PRIMARY KEY,
                    frete_id INTEGER NOT NULL REFERENCES fretes_freterequest(id) ON DELETE CASCADE,
                    endereco VARCHAR(255) NOT NULL,
                    cidade VARCHAR(100) NOT NULL,
                    estado VARCHAR(2) NOT NULL,
                    cep VARCHAR(10) NOT NULL,
                    volume INTEGER DEFAULT 1,
                    loja VARCHAR(100) DEFAULT '',
                    numero VARCHAR(20) DEFAULT '',
                    data_entrega TIMESTAMP WITH TIME ZONE,
                    observacao TEXT DEFAULT '',
                    anexo_destino VARCHAR(100) DEFAULT ''
                );
            """)
            print("✅ Tabela fretes_destino OK")
            
            # 6. CRIAR TABELA FRETES_COTACAOFRETE SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_cotacaofrete (
                    id SERIAL PRIMARY KEY,
                    frete_id INTEGER NOT NULL REFERENCES fretes_freterequest(id) ON DELETE CASCADE,
                    transportadora_id INTEGER NOT NULL REFERENCES fretes_transportadora(id) ON DELETE CASCADE,
                    valor_frete DECIMAL(10,2),
                    valor_pedagio DECIMAL(10,2),
                    valor_ajudante DECIMAL(10,2),
                    valor_total DECIMAL(10,2),
                    status VARCHAR(30) DEFAULT 'pendente',
                    data_cotacao TIMESTAMP WITH TIME ZONE,
                    observacoes_cotacao TEXT DEFAULT '',
                    motivo_rejeicao_transportadora TEXT DEFAULT '',
                    aprovador_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                    data_aprovacao TIMESTAMP WITH TIME ZONE,
                    observacoes_aprovacao TEXT DEFAULT '',
                    justificativa_rejeicao TEXT DEFAULT ''
                );
            """)
            print("✅ Tabela fretes_cotacaofrete OK")
            
            # 7. CRIAR TABELA FRETES_USERPROFILE SE NÃO EXISTIR
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS fretes_userprofile (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER NOT NULL REFERENCES auth_user(id) ON DELETE CASCADE,
                    tipo_acesso VARCHAR(20) DEFAULT 'limitado',
                    is_master BOOLEAN DEFAULT FALSE,
                    tipo_usuario VARCHAR(20) DEFAULT 'solicitante',
                    transportadora_id INTEGER REFERENCES fretes_transportadora(id) ON DELETE SET NULL
                );
            """)
            print("✅ Tabela fretes_userprofile OK")
            
            # 8. CRIAR ÍNDICES ÚNICOS
            cursor.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS fretes_userprofile_user_id_key 
                ON fretes_userprofile (user_id);
            """)
            print("✅ Índices únicos criados")
            
            # 9. CRIAR USUÁRIO ADMIN SE NÃO EXISTIR
            cursor.execute("SELECT id FROM auth_user WHERE username = 'admin';")
            admin_exists = cursor.fetchone()
            
            if not admin_exists:
                cursor.execute("""
                    INSERT INTO auth_user (username, email, password, is_staff, is_superuser, is_active, date_joined)
                    VALUES ('admin', 'admin@portal.com', 'pbkdf2_sha256$600000$dummy$dummy', TRUE, TRUE, TRUE, NOW());
                """)
                print("✅ Usuário admin criado")
                
                # Criar profile do admin
                cursor.execute("SELECT id FROM auth_user WHERE username = 'admin';")
                admin_id = cursor.fetchone()[0]
                
                cursor.execute("""
                    INSERT INTO fretes_userprofile (user_id, tipo_acesso, is_master, tipo_usuario)
                    VALUES (%s, 'completo', TRUE, 'master')
                    ON CONFLICT (user_id) DO UPDATE SET
                    tipo_acesso = 'completo',
                    is_master = TRUE,
                    tipo_usuario = 'master';
                """, [admin_id])
                print("✅ Profile do admin criado")
            else:
                print("✅ Usuário admin já existe")
            
            print("🎉 TODAS AS TABELAS VERIFICADAS E CRIADAS COM SUCESSO!")
                    
    except Exception as e:
        print(f"❌ Erro ao criar tabelas: {e}")
        import traceback
        print(f"Detalhes: {traceback.format_exc()}")

def home(request):
    """Página inicial com dashboard específico para cada tipo de usuário"""
    mensagem = request.GET.get('mensagem')
    
    # As tabelas são criadas automaticamente pelo middleware
    
    # Verificar se o usuário está logado
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = request.user.userprofile
        
        # Verificar tipo de usuário e renderizar template apropriado
        if user_profile.is_transportadora():
            return render(request, 'fretes/home_transportadora.html', {
                'mensagem': mensagem,
                'user_profile': user_profile
            })
        elif user_profile.is_gerente() or user_profile.is_usuario_master():
            return render(request, 'fretes/home_gerente.html', {
                'mensagem': mensagem,
                'user_profile': user_profile
            })
        else:
            # Usuário solicitante - usar template original com estatísticas
            fretes = FreteRequest.objects.select_related(
                'usuario', 'transportadora_selecionada'
            ).all().order_by('-data_criacao')
            
            # Estatísticas para o resumo rápido
            estatisticas = None
            try:
                # Total de fretes do usuário
                total_fretes = FreteRequest.objects.filter(usuario=request.user).count()
                
                # Fretes pendentes do usuário
                fretes_pendentes = FreteRequest.objects.filter(
                    usuario=request.user, 
                    status='pendente'
                ).count()
                
                # Fretes finalizados do usuário
                fretes_finalizados = FreteRequest.objects.filter(
                    usuario=request.user, 
                    status='finalizado'
                ).count()
                
                # Fretes deste mês do usuário
                hoje = timezone.now()
                inicio_mes = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                fretes_mes = FreteRequest.objects.filter(
                    usuario=request.user,
                    data_criacao__gte=inicio_mes
                ).count()
                
                estatisticas = {
                    'total_fretes': total_fretes,
                    'fretes_pendentes': fretes_pendentes,
                    'fretes_finalizados': fretes_finalizados,
                    'fretes_mes': fretes_mes,
                }
                
            except Exception as e:
                print(f"ERRO ao calcular estatísticas: {e}")
                estatisticas = {
                    'total_fretes': 0,
                    'fretes_pendentes': 0,
                    'fretes_finalizados': 0,
                    'fretes_mes': 0,
                }
            
            return render(request, 'fretes/home.html', {
                'fretes': fretes, 
                'mensagem': mensagem,
                'estatisticas': estatisticas
            })
            
    except UserProfile.DoesNotExist:
        # Se não tem perfil, usar template original
        fretes = FreteRequest.objects.select_related(
            'usuario', 'transportadora_selecionada'
        ).all().order_by('-data_criacao')
        
        return render(request, 'fretes/home.html', {
            'fretes': fretes, 
            'mensagem': mensagem,
            'estatisticas': None
        })




# Views para criação de fretes
@login_required(login_url='/login/')
def selecionar_origem(request):
    """Tela 1: Seleção de origem - OTIMIZADO com cache"""
    from django.core.cache import cache
    
    # Cache das lojas por 30 minutos (dados estáticos)
    cache_key = 'lojas_choices_sorted'
    lojas_choices = cache.get(cache_key)
    
    if lojas_choices is None:
        lojas_qs = Loja.objects.all().order_by('nome')
        
        def loja_numero(loja):
            try:
                # Extrair número do nome "Loja 18" -> 18
                if loja.nome.startswith('Loja '):
                    return int(loja.nome.replace('Loja ', ''))
                else:
                    return int(loja.nome)
            except (ValueError, TypeError):
                return 0
        
        lojas_list = sorted(lojas_qs, key=loja_numero)
        lojas_choices = [
            (str(loja.id), loja.nome, loja.latitude, loja.longitude) 
            for loja in lojas_list
        ]
        cache.set(cache_key, lojas_choices, 1800)  # 30 minutos
    
    if request.method == 'POST':
        loja_id = request.POST.get('origem')
        horario_coleta = request.POST.get('horario_coleta')
        tipo_veiculo = request.POST.get('tipo_veiculo')
        observacoes_origem = request.POST.get('observacoes_origem', '')
        anexo_origem = request.FILES.get('anexo_origem')
        
        # Novos campos para ajudante
        precisa_ajudante = request.POST.get('precisa_ajudante') == 'on'
        quantidade_ajudantes = request.POST.get('quantidade_ajudantes', '0')
        
        if loja_id and horario_coleta and tipo_veiculo:
            # Codificar as observações para URL
            import urllib.parse
            observacoes_encoded = urllib.parse.quote(observacoes_origem)
            
            # Se há anexo, salvar temporariamente na sessão
            if anexo_origem:
                # Validar arquivo
                if anexo_origem.size > 10 * 1024 * 1024:  # 10MB
                    return render(request, 'fretes/selecionar_origem.html', {
                        'lojas_choices': lojas_choices,
                        'erro': 'Arquivo muito grande. Tamanho máximo permitido: 10MB'
                    })
                
                allowed_extensions = ['.xlsx', '.xls', '.pdf']
                file_extension = os.path.splitext(anexo_origem.name)[1].lower()
                if file_extension not in allowed_extensions:
                    return render(request, 'fretes/selecionar_origem.html', {
                        'lojas_choices': lojas_choices,
                        'erro': 'Formato não permitido. Use apenas arquivos Excel (.xlsx, .xls) ou PDF (.pdf)'
                    })
                
                # Salvar arquivo temporariamente
                import tempfile
                import shutil
                temp_dir = tempfile.mkdtemp()
                temp_file_path = os.path.join(temp_dir, anexo_origem.name)
                with open(temp_file_path, 'wb') as temp_file:
                    for chunk in anexo_origem.chunks():
                        temp_file.write(chunk)
                
                # Armazenar informações na sessão
                request.session['anexo_origem_temp'] = {
                    'path': temp_file_path,
                    'name': anexo_origem.name,
                    'size': anexo_origem.size
                }
            
            # Codificar dados de ajudante para URL
            precisa_ajudante_encoded = '1' if precisa_ajudante else '0'
            quantidade_ajudantes_encoded = quantidade_ajudantes if precisa_ajudante else '0'
            
            return redirect(f"{reverse('selecionar_destino')}?origem_id={loja_id}&horario_coleta={horario_coleta}&tipo_veiculo={tipo_veiculo}&observacoes_origem={observacoes_encoded}&precisa_ajudante={precisa_ajudante_encoded}&quantidade_ajudantes={quantidade_ajudantes_encoded}")
    
    return render(request, 'fretes/selecionar_origem.html', {
        'lojas_choices': lojas_choices
    })


@login_required(login_url='/login/')
def selecionar_destino(request):
    """Tela 2: Seleção de destino estruturada"""
    origem_id = request.GET.get('origem_id') or request.POST.get('origem_id')
    horario_coleta = request.GET.get('horario_coleta') or request.POST.get('horario_coleta')
    tipo_veiculo = request.GET.get('tipo_veiculo') or request.POST.get('tipo_veiculo')
    observacoes_origem = request.GET.get('observacoes_origem') or request.POST.get('observacoes_origem', '')
    
    # Dados de ajudante
    precisa_ajudante = request.GET.get('precisa_ajudante', '0') == '1'
    quantidade_ajudantes = request.GET.get('quantidade_ajudantes', '0')
    
    # Decodificar as observações se vieram da URL
    if observacoes_origem and request.GET.get('observacoes_origem'):
        import urllib.parse
        observacoes_origem = urllib.parse.unquote(observacoes_origem)
    
    if not origem_id:
        return redirect('selecionar_origem')
    
    lojas_qs = Loja.objects.all()
    
    def loja_numero(loja):
        try:
            # Extrair número do nome "Loja 18" -> 18
            if loja.nome.startswith('Loja '):
                return int(loja.nome.replace('Loja ', ''))
            else:
                return int(loja.nome)
        except (ValueError, TypeError):
            return 0
    
    lojas_list = sorted(lojas_qs, key=loja_numero)
    lojas_choices = [
        (str(loja.id), loja.nome) 
        for loja in lojas_list 
        if str(loja.id) != str(origem_id)
    ]
    
    erro = None
    
    if request.method == 'POST':
        # Garantir que horario_coleta está disponível
        if not horario_coleta:
            horario_coleta = request.POST.get('horario_coleta') or request.GET.get('horario_coleta')
        
        destino_ids = request.POST.getlist('destino')
        descricao = request.POST.get('descricao', '')
        origem_loja = Loja.objects.filter(id=origem_id).first()
        
        print(f"DEBUG selecionar_destino - destino_ids recebidos: {destino_ids}")
        print(f"DEBUG selecionar_destino - origem_loja: {origem_loja}")
        
        if not origem_loja:
            erro = 'Origem inválida.'
        elif not destino_ids:
            erro = 'Selecione pelo menos um destino.'
        else:
            # Redirecionar para página de confirmação
            import urllib.parse
            
            # Preparar dados dos destinos
            volumes_data = []
            observacoes_data = []
            datas_entrega_data = []
            
            for loja_id in destino_ids:
                volume = request.POST.get(f'volume_{loja_id}', 1)
                observacao = request.POST.get(f'observacao_{loja_id}', '')
                data_entrega = request.POST.get(f'data_entrega_{loja_id}', '')
                
                volumes_data.append(f"{loja_id}:{volume}")
                observacoes_data.append(f"{loja_id}:{observacao}")
                datas_entrega_data.append(f"{loja_id}:{data_entrega}")
            
            # Buscar informações dos destinos para exibição
            destinos_info = []
            for loja_id in destino_ids:
                loja = Loja.objects.filter(id=loja_id).first()
                if loja:
                    destinos_info.append({
                        'id': loja_id,
                        'loja': loja.nome,
                        'endereco': loja.endereco,
                        'cidade': loja.municipio,
                        'estado': loja.estado
                    })
            
            return render(request, 'fretes/confirmar_frete.html', {
                'origem': origem_loja,
                'origem_id': origem_id,
                'horario_coleta': horario_coleta,
                'tipo_veiculo': tipo_veiculo,
                'observacoes_origem': observacoes_origem,
                'precisa_ajudante': precisa_ajudante,
                'quantidade_ajudantes': quantidade_ajudantes,
                'destino_ids': ','.join(destino_ids),
                'volumes': ','.join(volumes_data),
                'observacoes_destinos': ','.join(observacoes_data),
                'datas_entrega': ','.join(datas_entrega_data),
                'destinos_info': destinos_info,
                'destinos_count': len(destino_ids)
            })
    
    return render(request, 'fretes/selecionar_destino.html', {
        'lojas_choices': lojas_choices,
        'origem_id': origem_id,
        'horario_coleta': horario_coleta,
        'tipo_veiculo': tipo_veiculo,
        'observacoes_origem': observacoes_origem,
        'precisa_ajudante': precisa_ajudante,
        'quantidade_ajudantes': quantidade_ajudantes,
        'erro': erro
    })


@login_required(login_url='/login/')
def confirmar_frete(request):
    """Tela 3: Confirmação final e criação do frete"""
    if request.method == 'POST':
        # Processar dados da confirmação
        origem_id = request.POST.get('origem_id')
        horario_coleta = request.POST.get('horario_coleta')
        tipo_veiculo = request.POST.get('tipo_veiculo')
        observacoes_origem = request.POST.get('observacoes_origem', '')
        destino_ids_str = request.POST.get('destino_ids', '')
        destino_ids = destino_ids_str.split(',') if destino_ids_str else []
        volumes_data = request.POST.get('volumes', '').split(',')
        observacoes_destinos = request.POST.get('observacoes_destinos', '').split(',')
        datas_entrega_data = request.POST.get('datas_entrega', '').split(',')
        
        # Novos campos
        nota_fiscal_emitida = request.POST.get('nota_fiscal_emitida') == 'true'
        anexo_nota_fiscal = request.FILES.get('anexo_nota_fiscal')
        quem_paga_frete = request.POST.get('quem_paga_frete')
        
        # Dados de ajudante
        precisa_ajudante = request.POST.get('precisa_ajudante') == 'on'
        quantidade_ajudantes = request.POST.get('quantidade_ajudantes', '0')
        
        print(f"DEBUG - Dados recebidos: origem_id={origem_id}, destino_ids_str='{destino_ids_str}', destino_ids={destino_ids}, quem_paga_frete={quem_paga_frete}")
        
        # Validar dados obrigatórios
        if not origem_id or not horario_coleta or not tipo_veiculo:
            print(f"DEBUG - Dados obrigatórios básicos: origem_id={origem_id}, horario_coleta={horario_coleta}, tipo_veiculo={tipo_veiculo}")
            # Redirecionar para seleção de origem se dados básicos estão faltando
            return redirect('selecionar_origem')
        
        # Validar quem_paga_frete separadamente
        if not quem_paga_frete:
            print(f"DEBUG - quem_paga_frete não selecionado: {quem_paga_frete}")
            # Redirecionar para seleção de destino se quem paga não foi selecionado
            return redirect('selecionar_destino')
        
        # Validar destino_ids
        if not destino_ids or (len(destino_ids) == 1 and not destino_ids[0]):
            print(f"DEBUG - Erro: destino_ids vazio ou inválido: {destino_ids}")
            # Redirecionar para seleção de destino se não há destinos
            return redirect('selecionar_destino')
        
        # Buscar origem
        origem_loja = Loja.objects.filter(id=origem_id).first()
        if not origem_loja:
            return render(request, 'fretes/confirmar_frete.html', {
                'erro': 'Origem inválida.'
            })
        
        # Criar o frete
        frete = FreteRequest.objects.create(
            usuario=request.user,
            origem=origem_loja,
            horario_coleta=horario_coleta,
            tipo_veiculo=tipo_veiculo,
            observacoes_origem=observacoes_origem,
            nota_fiscal_emitida=nota_fiscal_emitida,
            quem_paga_frete=quem_paga_frete,
            precisa_ajudante=precisa_ajudante,
            quantidade_ajudantes=int(quantidade_ajudantes) if quantidade_ajudantes.isdigit() else 0
        )
        
        # Processar anexo da origem se existir
        if 'anexo_origem_temp' in request.session:
            anexo_info = request.session['anexo_origem_temp']
            try:
                from django.core.files import File
                with open(anexo_info['path'], 'rb') as temp_file:
                    frete.anexo_origem.save(anexo_info['name'], File(temp_file), save=True)
                
                print(f"DEBUG: Anexo origem salvo - Caminho: {frete.anexo_origem.path}")
                print(f"DEBUG: Anexo origem URL: {frete.anexo_origem.url}")
                print(f"DEBUG: Anexo origem existe: {os.path.exists(frete.anexo_origem.path)}")
                
                os.unlink(anexo_info['path'])
                os.rmdir(os.path.dirname(anexo_info['path']))
                del request.session['anexo_origem_temp']
            except Exception as e:
                print(f"Erro ao processar anexo da origem: {e}")
        
        # Processar anexo da nota fiscal se existir
        if anexo_nota_fiscal:
            # Validar arquivo
            if anexo_nota_fiscal.size <= 10 * 1024 * 1024:  # 10MB
                allowed_extensions = ['.pdf', '.xlsx', '.xls']
                file_extension = os.path.splitext(anexo_nota_fiscal.name)[1].lower()
                if file_extension in allowed_extensions:
                    frete.anexo_nota_fiscal = anexo_nota_fiscal
                    frete.save()
                    
                    print(f"DEBUG: Anexo nota fiscal salvo - Caminho: {frete.anexo_nota_fiscal.path}")
                    print(f"DEBUG: Anexo nota fiscal URL: {frete.anexo_nota_fiscal.url}")
                    print(f"DEBUG: Anexo nota fiscal existe: {os.path.exists(frete.anexo_nota_fiscal.path)}")
        
        # Criar os destinos
        for loja_id in destino_ids:
            if not loja_id:
                continue
                
            loja = Loja.objects.filter(id=loja_id).first()
            if loja:
                # Buscar volume
                volume = 1
                for vol_data in volumes_data:
                    if vol_data.startswith(f"{loja_id}:"):
                        try:
                            volume = int(vol_data.split(':', 1)[1])
                            if volume < 1:
                                volume = 1
                        except (ValueError, TypeError):
                            volume = 1
                        break
                
                # Buscar observação
                observacao = ''
                for obs_data in observacoes_destinos:
                    if obs_data.startswith(f"{loja_id}:"):
                        observacao = obs_data.split(':', 1)[1]
                        break
                
                # Buscar data de entrega
                data_entrega = None
                for data_data in datas_entrega_data:
                    if data_data.startswith(f"{loja_id}:"):
                        data_str = data_data.split(':', 1)[1]
                        if data_str:
                            try:
                                from django.utils.dateparse import parse_datetime
                                data_entrega = parse_datetime(data_str)
                            except (ValueError, TypeError):
                                data_entrega = None
                        break
                
                destino = Destino.objects.create(
                    frete=frete,
                    loja=loja.nome,
                    endereco=loja.endereco,
                    numero=loja.numero,
                    cidade=loja.municipio,
                    estado=loja.estado,
                    cep=loja.cep,
                    volume=volume,
                    data_entrega=data_entrega,
                    observacao=observacao,
                )
        
        return redirect(f"{reverse('home')}?mensagem=Frete+enviado+com+sucesso")
    
    # Se não for POST, redirecionar para seleção de origem
    return redirect('selecionar_origem')


# Views para gerenciar fretes
@login_required(login_url='/login/')
def meus_fretes(request):
    """Lista de fretes do usuário logado (ou todos se for master)"""
    status = request.GET.get('status')
    search = request.GET.get('search', '').strip()
    
    # Verificar se é usuário master
    try:
        user_profile = request.user.userprofile
        if user_profile.is_usuario_master():
            # Master vê todos os fretes - OTIMIZADO com select_related e prefetch_related
            qs = FreteRequest.objects.select_related(
                'usuario', 'origem', 'transportadora_selecionada', 'aprovador'
            ).prefetch_related(
                'destinos', 'cotacoes__transportadora', 'agendamento__tracking'
            )
        elif user_profile.is_transportadora() and user_profile.transportadora:
            # Transportadora vê apenas fretes direcionados para ela - OTIMIZADO
            qs = FreteRequest.objects.filter(
                cotacaofrete__transportadora=user_profile.transportadora
            ).select_related(
                'usuario', 'origem', 'transportadora_selecionada', 'aprovador'
            ).prefetch_related(
                'destinos', 'cotacoes__transportadora', 'agendamento__tracking'
            ).distinct()
        else:
            # Usuário normal vê apenas seus fretes - OTIMIZADO
            qs = FreteRequest.objects.filter(usuario=request.user).select_related(
                'usuario', 'origem', 'transportadora_selecionada', 'aprovador'
            ).prefetch_related(
                'destinos', 'cotacoes__transportadora', 'agendamento__tracking'
            )
    except UserProfile.DoesNotExist:
        # Se não tem perfil, vê apenas seus fretes - OTIMIZADO
        qs = FreteRequest.objects.filter(usuario=request.user).select_related(
            'usuario', 'origem', 'transportadora_selecionada', 'aprovador'
        ).prefetch_related(
            'destinos', 'cotacoes__transportadora', 'agendamento__tracking'
        )
    
    # Filtro por status
    if status in ['pendente', 'aguardando_aprovacao', 'aprovado', 'rejeitado', 'cotacao_enviada', 'cotacao_recebida', 'cotacao_aprovada', 'agendado', 'em_transito', 'entregue', 'finalizado']:
        qs = qs.filter(status=status)
    
    # Filtro por pesquisa - OTIMIZADO com índices
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(id__icontains=search) |
            Q(origem__nome__icontains=search) |
            Q(origem__municipio__icontains=search) |
            Q(descricao__icontains=search) |
            Q(destinos__loja__icontains=search) |
            Q(destinos__cidade__icontains=search)
        ).distinct()
    
    fretes = qs.order_by('-data_criacao')
    return render(request, 'fretes/meus_fretes.html', {
        'fretes': fretes, 
        'status_selected': status,
        'search_query': search
    })


@login_required(login_url='/login/')
def frete_detalhe(request, frete_id):
    """Detalhes de um frete específico"""
    # OTIMIZADO: Carregar todas as relações necessárias em uma única consulta
    frete = get_object_or_404(
        FreteRequest.objects.select_related(
            'usuario', 'origem', 'transportadora_selecionada', 'aprovador', 'usuario_cancelamento'
        ).prefetch_related(
            'destinos', 'cotacoes__transportadora', 'cotacoes__aprovador', 'agendamento__tracking'
        ), 
        id=frete_id
    )
    
    # Verificar se o usuário pode ver detalhes do frete
    try:
        user_profile = request.user.userprofile
        # Permitir acesso se:
        # 1. É o dono do frete (sempre pode ver seus próprios fretes)
        # 2. É usuário master/gerente (pode ver todos os fretes)
        # 3. Tem acesso completo (pode ver todos os fretes)
        # 4. É transportadora e o frete foi direcionado para ela
        pode_ver = (
            frete.usuario == request.user or 
            user_profile.is_usuario_master() or 
            user_profile.pode_ver_detalhes_frete() or
            (user_profile.is_transportadora() and user_profile.transportadora and 
             CotacaoFrete.objects.filter(frete=frete, transportadora=user_profile.transportadora).exists())
        )
        
        if not pode_ver:
            messages.error(request, 'Você não tem permissão para visualizar detalhes deste frete.')
            return redirect('meus_fretes')
    except UserProfile.DoesNotExist:
        # Se não tem perfil, só pode ver seus próprios fretes
        if frete.usuario != request.user:
            messages.error(request, 'Perfil de usuário não encontrado.')
            return redirect('meus_fretes')
    
    destinos = frete.destinos.all()
    transportadoras = Transportadora.objects.all()
    
    # Buscar cotação aprovada para mostrar a transportadora aprovada
    cotacao_aprovada = None
    transportadora_aprovada = None
    if frete.status == 'cotacao_aprovada':
        cotacao_aprovada = CotacaoFrete.objects.filter(
            frete=frete, 
            status='aprovada'
        ).select_related('transportadora').first()
        if cotacao_aprovada:
            transportadora_aprovada = cotacao_aprovada.transportadora
    
    # Buscar informações de agendamento se existir
    agendamento = getattr(frete, 'agendamento', None)
    tracking_history = []
    if agendamento:
        tracking_history = agendamento.tracking.all().order_by('-data_atualizacao')
    
    if request.method == 'POST':
        transportadora_id = request.POST.get('transportadora_id')
        if transportadora_id:
            transportadora = Transportadora.objects.filter(id=transportadora_id).first()
            if transportadora:
                frete.transportadora_selecionada = transportadora
                frete.save()
    
    return render(request, 'fretes/frete_detalhe.html', {
        'frete': frete,
        'destinos': destinos,
        'transportadoras': transportadoras,
        'cotacao_aprovada': cotacao_aprovada,
        'transportadora_aprovada': transportadora_aprovada,
        'agendamento': agendamento,
        'tracking_history': tracking_history
    })


@login_required(login_url='/login/')
def editar_frete(request, frete_id):
    """Editar um frete existente"""
    frete = get_object_or_404(FreteRequest, id=frete_id)
    
    # Verificar se o usuário pode editar fretes
    try:
        user_profile = request.user.userprofile
        # Permitir edição apenas se:
        # 1. É usuário master/gerente (pode editar todos os fretes)
        # 2. Tem acesso completo (pode editar todos os fretes)
        if not user_profile.is_usuario_master() and not user_profile.pode_editar_fretes():
            messages.error(request, 'Você não tem permissão para editar fretes.')
            return redirect('meus_fretes')
    except UserProfile.DoesNotExist:
        # Se não tem perfil, não pode editar
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('meus_fretes')
    
    # Verificar se o frete pode ser editado (apenas pendentes)
    if frete.status != 'pendente':
        return redirect('frete_detalhe', frete_id=frete_id)
    
    # Buscar todas as lojas para os selects
    lojas_qs = Loja.objects.all()
    
    def loja_numero(loja):
        try:
            # Extrair número do nome "Loja 18" -> 18
            if loja.nome.startswith('Loja '):
                return int(loja.nome.replace('Loja ', ''))
            else:
                return int(loja.nome)
        except (ValueError, TypeError):
            return 0
    
    lojas_list = sorted(lojas_qs, key=loja_numero)
    lojas_choices = [
        (str(loja.id), loja.nome, loja.latitude, loja.longitude) 
        for loja in lojas_list
    ]
    
    # Buscar destinos existentes
    destinos_existentes = frete.destinos.all()
    
    if request.method == 'POST':
        # Processar dados do formulário
        origem_id = request.POST.get('origem')
        horario_coleta = request.POST.get('horario_coleta')
        tipo_veiculo = request.POST.get('tipo_veiculo')
        observacoes_origem = request.POST.get('observacoes_origem', '')
        anexo_origem = request.FILES.get('anexo_origem')
        
        # Validar dados obrigatórios
        if not origem_id or not horario_coleta or not tipo_veiculo:
            return render(request, 'fretes/editar_frete.html', {
                'frete': frete,
                'lojas_choices': lojas_choices,
                'destinos_existentes': destinos_existentes,
                'erro': 'Origem, horário de coleta e tipo de veículo são obrigatórios.'
            })
        
        # Atualizar dados básicos do frete
        origem_loja = Loja.objects.filter(id=origem_id).first()
        if origem_loja:
            frete.origem = origem_loja
            frete.horario_coleta = horario_coleta
            frete.tipo_veiculo = tipo_veiculo
            frete.observacoes_origem = observacoes_origem
            
            # Processar novo anexo da origem se fornecido
            if anexo_origem:
                # Validar arquivo
                if anexo_origem.size > 10 * 1024 * 1024:  # 10MB
                    return render(request, 'fretes/editar_frete.html', {
                        'frete': frete,
                        'lojas_choices': lojas_choices,
                        'destinos_existentes': destinos_existentes,
                        'erro': 'Arquivo muito grande. Tamanho máximo permitido: 10MB'
                    })
                
                allowed_extensions = ['.xlsx', '.xls', '.pdf']
                file_extension = os.path.splitext(anexo_origem.name)[1].lower()
                if file_extension not in allowed_extensions:
                    return render(request, 'fretes/editar_frete.html', {
                        'frete': frete,
                        'lojas_choices': lojas_choices,
                        'destinos_existentes': destinos_existentes,
                        'erro': 'Formato não permitido. Use apenas arquivos Excel (.xlsx, .xls) ou PDF (.pdf)'
                    })
                
                # Deletar anexo antigo se existir
                if frete.anexo_origem:
                    try:
                        frete.anexo_origem.delete(save=False)
                    except:
                        pass
                
                # Salvar novo anexo
                frete.anexo_origem = anexo_origem
            
            frete.save()
            
            # Processar destinos
            destino_ids = request.POST.getlist('destino')
            if destino_ids:
                # Deletar destinos existentes
                frete.destinos.all().delete()
                
                # Criar novos destinos
                for loja_id in destino_ids:
                    loja = Loja.objects.filter(id=loja_id).first()
                    if loja:
                        volume = request.POST.get(f'volume_{loja_id}', 1)
                        try:
                            volume = int(volume)
                            if volume < 1:
                                raise ValueError
                        except (ValueError, TypeError):
                            volume = 1
                        
                        observacao = request.POST.get(f'observacao_{loja_id}', '')
                        data_entrega_str = request.POST.get(f'data_entrega_{loja_id}', '')
                        
                        # Processar data de entrega
                        data_entrega = None
                        if data_entrega_str:
                            try:
                                from django.utils.dateparse import parse_datetime
                                data_entrega = parse_datetime(data_entrega_str)
                            except (ValueError, TypeError):
                                data_entrega = None
                        
                        destino = Destino.objects.create(
                            frete=frete,
                            loja=loja.nome,
                            endereco=loja.endereco,
                            numero=loja.numero,
                            cidade=loja.municipio,
                            estado=loja.estado,
                            cep=loja.cep,
                            volume=volume,
                            data_entrega=data_entrega,
                            observacao=observacao,
                        )
                        
                        # Processar anexo do destino se existir
                        anexo_destino = request.FILES.get(f'anexo_destino_{loja_id}')
                        if anexo_destino:
                            # Validar arquivo
                            if anexo_destino.size > 10 * 1024 * 1024:  # 10MB
                                continue  # Pular este destino se arquivo for muito grande
                            
                            allowed_extensions = ['.xlsx', '.xls', '.pdf']
                            file_extension = os.path.splitext(anexo_destino.name)[1].lower()
                            if file_extension in allowed_extensions:
                                destino.anexo_destino = anexo_destino
                                destino.save()
            
            return redirect('frete_detalhe', frete_id=frete_id)
    
    return render(request, 'fretes/editar_frete.html', {
        'frete': frete,
        'lojas_choices': lojas_choices,
        'destinos_existentes': destinos_existentes
    })


@login_required(login_url='/login/')
@require_POST
def atualizar_status_frete(request, frete_id):
    """Atualizar status de um frete via AJAX"""
    frete = get_object_or_404(FreteRequest, id=frete_id)
    status = request.POST.get('status')
    
    if status in ['pendente', 'aguardando_aprovacao', 'aprovado', 'rejeitado', 'cotacao_enviada', 'finalizado']:
        frete.status = status
        frete.save()
        return JsonResponse({'success': True, 'status': status})
    
    return JsonResponse({
        'success': False, 
        'error': 'Status inválido'
    }, status=400)


# Views de API/AJAX
@login_required(login_url='/login/')
def loja_info(request, loja_id):
    """API para obter informações de uma loja"""
    loja = Loja.objects.filter(id=loja_id).first()
    
    if loja:
        # Formatar CEP removendo .0 se existir
        cep_formatted = str(loja.cep) if loja.cep else ''
        if cep_formatted.endswith('.0'):
            cep_formatted = cep_formatted[:-2]
        
        data = {
            'nome': loja.nome,
            'endereco': loja.endereco,
            'numero': loja.numero,
            'municipio': loja.municipio,
            'estado': loja.estado,
            'cep': cep_formatted,
            'latitude': str(loja.latitude) if loja.latitude else '',
            'longitude': str(loja.longitude) if loja.longitude else '',
        }
        return JsonResponse(data)
    
    return JsonResponse({'error': 'Loja não encontrada'}, status=404)


# Views para aprovação de fretes
@login_required(login_url='/login/')
def fretes_para_aprovacao(request):
    """Lista de fretes pendentes para direcionamento"""
    # Verificar se o usuário pode aprovar fretes
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_usuario_master() or user_profile.is_gerente()):
            messages.error(request, 'Você não tem permissão para aprovar fretes.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar fretes pendentes - OTIMIZADO
    fretes_pendentes = FreteRequest.objects.filter(
        status='pendente'
    ).select_related(
        'usuario', 'origem', 'transportadora_selecionada'
    ).prefetch_related(
        'destinos', 'cotacoes__transportadora'
    ).order_by('-data_criacao')
    
    # Buscar transportadoras disponíveis
    transportadoras = Transportadora.objects.all().order_by('nome')
    
    return render(request, 'fretes/fretes_para_aprovacao.html', {
        'fretes_pendentes': fretes_pendentes,
        'transportadoras': transportadoras
    })


@login_required(login_url='/login/')
def direcionar_frete_ajax(request):
    """Direciona um frete para múltiplas transportadoras via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)
    
    # Verificar permissões
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_usuario_master() or user_profile.is_gerente()):
            return JsonResponse({'error': 'Sem permissão'}, status=403)
    except UserProfile.DoesNotExist:
        return JsonResponse({'error': 'Perfil não encontrado'}, status=403)
    
    frete_id = request.POST.get('frete_id')
    transportadoras_ids = request.POST.getlist('transportadoras_ids[]')
    
    if not frete_id or not transportadoras_ids:
        return JsonResponse({'error': 'Dados incompletos'}, status=400)
    
    try:
        frete = FreteRequest.objects.get(id=frete_id, status='pendente')
        transportadoras = Transportadora.objects.filter(
            id__in=transportadoras_ids
        )
        
        if not transportadoras.exists():
            return JsonResponse({'error': 'Nenhuma transportadora válida encontrada'}, status=404)
        
        # Criar cotações para cada transportadora
        cotacoes_criadas = []
        for transportadora in transportadoras:
            # Verificar se já existe cotação para esta transportadora
            cotacao_existente = CotacaoFrete.objects.filter(
                frete=frete, 
                transportadora=transportadora
            ).first()
            
            if not cotacao_existente:
                cotacao = CotacaoFrete.objects.create(
                    frete=frete,
                    transportadora=transportadora,
                    status='pendente'
                )
                cotacoes_criadas.append(transportadora.nome)
        
        # Atualizar status do frete
        frete.status = 'cotacao_enviada'
        frete.save()
        
        if cotacoes_criadas:
            return JsonResponse({
                'success': True,
                'message': f'Frete #{frete.id} direcionado para {len(cotacoes_criadas)} transportadora(s): {", ".join(cotacoes_criadas)}'
            })
        else:
            return JsonResponse({
                'success': True,
                'message': f'Frete #{frete.id} já estava direcionado para as transportadoras selecionadas'
            })
        
    except FreteRequest.DoesNotExist:
        return JsonResponse({'error': 'Frete não encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='/login/')
def fretes_para_cotacao(request):
    """Lista de fretes direcionados para cotação pela transportadora"""
    # Verificar se o usuário é transportadora ou master
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_transportadora() or user_profile.is_usuario_master()):
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar cotações pendentes
    if user_profile.is_usuario_master():
        # Master vê todas as cotações pendentes
        cotacoes_pendentes = CotacaoFrete.objects.filter(
            status='pendente'
        ).select_related('frete__usuario', 'frete__origem', 'transportadora').prefetch_related('frete__destinos').order_by('-frete__data_criacao')
    else:
        # Transportadora vê apenas suas cotações
        if user_profile.transportadora:
            cotacoes_pendentes = CotacaoFrete.objects.filter(
                transportadora=user_profile.transportadora,
                status='pendente'
            ).select_related('frete__usuario', 'frete__origem').prefetch_related('frete__destinos').order_by('-frete__data_criacao')
        else:
            cotacoes_pendentes = CotacaoFrete.objects.none()
    
    return render(request, 'fretes/fretes_para_cotacao.html', {
        'cotacoes_pendentes': cotacoes_pendentes
    })


@login_required(login_url='/login/')
def cotar_frete(request, cotacao_id):
    """Tela de cotação de um frete específico pela transportadora"""
    # Verificar se o usuário é transportadora ou master
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_transportadora() or user_profile.is_usuario_master()):
            messages.error(request, 'Você não tem permissão para acessar esta área.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar a cotação
    try:
        if user_profile.is_usuario_master():
            # Master pode ver qualquer cotação
            cotacao = CotacaoFrete.objects.get(
                id=cotacao_id,
                status='pendente'
            )
        else:
            # Transportadora vê apenas suas cotações
            if user_profile.transportadora:
                cotacao = CotacaoFrete.objects.get(
                    id=cotacao_id,
                    transportadora=user_profile.transportadora,
                    status='pendente'
                )
            else:
                messages.error(request, 'Transportadora não vinculada ao seu usuário.')
                return redirect('home')
        frete = cotacao.frete
    except CotacaoFrete.DoesNotExist:
        messages.error(request, 'Cotação não encontrada ou não direcionada para você.')
        return redirect('fretes_para_cotacao')
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'cotar':
            # Processar cotação
            valor_frete = request.POST.get('valor_frete')
            valor_pedagio = request.POST.get('valor_pedagio')
            valor_ajudante = request.POST.get('valor_ajudante')
            centro_custo = request.POST.get('centro_custo')
            observacoes = request.POST.get('observacoes_cotacao')
            
            if not valor_frete:
                messages.error(request, 'Valor do frete é obrigatório.')
                return render(request, 'fretes/cotar_frete.html', {'cotacao': cotacao, 'frete': frete})
            
            try:
                cotacao.valor_frete = float(valor_frete) if valor_frete else None
                cotacao.valor_pedagio = float(valor_pedagio) if valor_pedagio else None
                cotacao.valor_ajudante = float(valor_ajudante) if valor_ajudante else None
                cotacao.observacoes_cotacao = observacoes
                cotacao.data_cotacao = timezone.now()
                cotacao.status = 'enviada'
                cotacao.save()
                
                # Atualizar o frete com centro de custo se fornecido
                if centro_custo:
                    frete.centro_custo = centro_custo
                    frete.save()
                
                # Atualizar status do frete para "Cotação Recebida"
                frete.status = 'cotacao_recebida'
                frete.data_cotacao = timezone.now()
                frete.save()
                
                messages.success(request, 'Cotação enviada com sucesso!')
                return redirect('fretes_para_cotacao')
                
            except ValueError:
                messages.error(request, 'Valores inválidos. Use apenas números.')
                return render(request, 'fretes/cotar_frete.html', {'cotacao': cotacao, 'frete': frete})
        
        elif acao == 'rejeitar':
            # Processar rejeição
            motivo = request.POST.get('motivo_rejeicao_transportadora')
            
            if not motivo:
                messages.error(request, 'Motivo da rejeição é obrigatório.')
                return render(request, 'fretes/cotar_frete.html', {'cotacao': cotacao, 'frete': frete})
            
            cotacao.motivo_rejeicao_transportadora = motivo
            cotacao.status = 'rejeitada_transportadora'
            cotacao.save()
            
            messages.success(request, 'Frete rejeitado com sucesso.')
            return redirect('fretes_para_cotacao')
    
    return render(request, 'fretes/cotar_frete.html', {'cotacao': cotacao, 'frete': frete})


@login_required(login_url='/login/')
def aprovar_frete(request, frete_id):
    """Tela de aprovação de um frete específico"""
    frete = get_object_or_404(FreteRequest, id=frete_id)
    
    # Verificar se o usuário pode aprovar fretes
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_usuario_master() or user_profile.is_gerente()):
            messages.error(request, 'Você não tem permissão para aprovar fretes.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Verificar se o frete está pendente
    if frete.status != 'pendente':
        messages.error(request, 'Este frete não está pendente para direcionamento.')
        return redirect('fretes_para_aprovacao')
    
    destinos = frete.destinos.all()
    
    # Buscar transportadoras disponíveis
    transportadoras = User.objects.filter(userprofile__tipo_usuario='transportadora')
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'direcionar':
            transportadora_id = request.POST.get('transportadora')
            observacoes = request.POST.get('observacoes_aprovacao', '').strip()
            
            if not transportadora_id:
                messages.error(request, 'Selecione uma transportadora.')
                return render(request, 'fretes/aprovar_frete.html', {
                    'frete': frete,
                    'destinos': destinos,
                    'transportadoras': transportadoras
                })
            
            try:
                transportadora = User.objects.get(id=transportadora_id)
                # Direcionar o frete para a transportadora
                frete.status = 'cotacao_enviada'
                frete.transportadora = transportadora
                frete.observacoes_aprovacao = observacoes
                frete.aprovador = request.user
                frete.data_aprovacao = timezone.now()
                frete.save()
                
                messages.success(request, f'Frete #{frete.id} direcionado para {transportadora.username} com sucesso!')
                return redirect('fretes_para_aprovacao')
                
            except User.DoesNotExist:
                messages.error(request, 'Transportadora não encontrada.')
                return render(request, 'fretes/aprovar_frete.html', {
                    'frete': frete,
                    'destinos': destinos,
                    'transportadoras': transportadoras
                })
            
        elif acao == 'rejeitar':
            justificativa = request.POST.get('justificativa_rejeicao', '').strip()
            
            if not justificativa:
                messages.error(request, 'Justificativa é obrigatória para rejeição.')
                return render(request, 'fretes/aprovar_frete.html', {
                    'frete': frete,
                    'destinos': destinos
                })
            
            # Rejeitar o frete
            frete.status = 'rejeitado'
            frete.justificativa_rejeicao = justificativa
            frete.aprovador = request.user
            frete.data_aprovacao = timezone.now()
            frete.save()
            
            messages.success(request, f'Frete #{frete.id} rejeitado.')
            return redirect('fretes_para_aprovacao')
    
    return render(request, 'fretes/aprovar_frete.html', {
        'frete': frete,
        'destinos': destinos,
        'transportadoras': transportadoras
    })


@login_required(login_url='/login/')
def cotacoes_recebidas(request):
    """Lista de cotações recebidas das transportadoras"""
    # Verificar se o usuário pode aprovar cotações
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_usuario_master() or user_profile.is_gerente()):
            messages.error(request, 'Você não tem permissão para aprovar cotações.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar cotações recebidas
    cotacoes_recebidas = CotacaoFrete.objects.filter(
        status='enviada'
    ).select_related('frete__usuario', 'frete__origem', 'transportadora').prefetch_related('frete__destinos').order_by('-data_cotacao')
    
    return render(request, 'fretes/cotacoes_recebidas.html', {
        'cotacoes_recebidas': cotacoes_recebidas
    })


@login_required(login_url='/login/')
def aprovar_cotacao(request, cotacao_id):
    """Tela de aprovação de uma cotação específica"""
    # Verificar se o usuário pode aprovar cotações
    try:
        user_profile = request.user.userprofile
        if not (user_profile.is_usuario_master() or user_profile.is_gerente()):
            messages.error(request, 'Você não tem permissão para aprovar cotações.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar a cotação
    try:
        cotacao = CotacaoFrete.objects.get(id=cotacao_id, status='enviada')
        frete = cotacao.frete
    except CotacaoFrete.DoesNotExist:
        messages.error(request, 'Cotação não encontrada ou já processada.')
        return redirect('cotacoes_recebidas')
    
    destinos = frete.destinos.all()
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'aprovar':
            centro_custo = request.POST.get('centro_custo', '').strip()
            observacoes = request.POST.get('observacoes_aprovacao', '').strip()
            
            if not centro_custo:
                messages.error(request, 'Centro de custo é obrigatório para aprovação.')
                return render(request, 'fretes/aprovar_cotacao.html', {
                    'cotacao': cotacao,
                    'frete': frete,
                    'destinos': destinos
                })
            
            # Aprovar a cotação
            cotacao.status = 'aprovada'
            cotacao.observacoes_aprovacao = observacoes
            cotacao.aprovador = request.user
            cotacao.data_aprovacao = timezone.now()
            cotacao.save()
            
            # Atualizar o frete com a cotação aprovada
            frete.status = 'cotacao_aprovada'
            frete.centro_custo = centro_custo
            frete.transportadora_selecionada = cotacao.transportadora
            frete.valor_frete = cotacao.valor_frete
            frete.valor_pedagio = cotacao.valor_pedagio
            frete.valor_ajudante = cotacao.valor_ajudante
            frete.valor_total = cotacao.valor_total
            frete.observacoes_cotacao = cotacao.observacoes_cotacao
            frete.data_cotacao = cotacao.data_cotacao
            frete.observacoes_aprovacao = observacoes
            frete.aprovador = request.user
            frete.data_aprovacao = timezone.now()
            frete.save()
            
            messages.success(request, f'Cotação #{cotacao.id} aprovada com sucesso!')
            return redirect('cotacoes_recebidas')
        
        elif acao == 'rejeitar':
            justificativa = request.POST.get('justificativa_rejeicao', '').strip()
            
            if not justificativa:
                messages.error(request, 'Justificativa é obrigatória para rejeição.')
                return render(request, 'fretes/aprovar_cotacao.html', {
                    'cotacao': cotacao,
                    'frete': frete,
                    'destinos': destinos
                })
            
            # Rejeitar a cotação
            cotacao.status = 'rejeitada'
            cotacao.justificativa_rejeicao = justificativa
            cotacao.aprovador = request.user
            cotacao.data_aprovacao = timezone.now()
            cotacao.save()
            
            # Voltar o frete para pendente para poder ser direcionado para outra transportadora
            frete.status = 'pendente'
            frete.save()
            
            messages.success(request, f'Cotação #{cotacao.id} rejeitada.')
            return redirect('cotacoes_recebidas')
    
    return render(request, 'fretes/aprovar_cotacao.html', {
        'cotacao': cotacao,
        'frete': frete,
        'destinos': destinos
    })


# Views para gerenciamento de usuários
@login_required(login_url='/login/')
def gerenciar_usuarios(request):
    """Lista de usuários para gerenciamento pelo master"""
    # Verificar se o usuário é master (apenas master, não gerente)
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_master or user_profile.tipo_usuario != 'master':
            messages.error(request, 'Você não tem permissão para gerenciar usuários.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar todos os usuários com seus perfis
    usuarios = User.objects.select_related('userprofile').all().order_by('-date_joined')
    
    return render(request, 'fretes/gerenciar_usuarios.html', {
        'usuarios': usuarios
    })


@login_required(login_url='/login/')
def criar_usuario(request):
    """Criar novo usuário"""
    # Verificar se o usuário é master
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para criar usuários.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        password = request.POST.get('password', '').strip()
        tipo_usuario = request.POST.get('tipo_usuario', 'solicitante')
        
        # Validações
        if not username or not password:
            messages.error(request, 'Username e senha são obrigatórios.')
            return render(request, 'fretes/criar_usuario.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username já existe.')
            return render(request, 'fretes/criar_usuario.html')
        
        if email and User.objects.filter(email=email).exists():
            messages.error(request, 'Email já existe.')
            return render(request, 'fretes/criar_usuario.html')
        
        try:
            # Criar usuário
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            
            # Atualizar perfil
            profile = user.userprofile
            profile.tipo_usuario = tipo_usuario
            if tipo_usuario == 'master':
                profile.is_master = True
                profile.tipo_acesso = 'completo'
            elif tipo_usuario == 'gerente':
                profile.is_master = False
                profile.tipo_acesso = 'completo'
            elif tipo_usuario == 'transportadora':
                profile.tipo_acesso = 'completo'
            else:  # solicitante
                profile.tipo_acesso = 'limitado'
            profile.save()
            
            messages.success(request, f'Usuário {username} criado com sucesso!')
            return redirect('gerenciar_usuarios')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar usuário: {str(e)}')
            return render(request, 'fretes/criar_usuario.html')
    
    return render(request, 'fretes/criar_usuario.html')


@login_required(login_url='/login/')
def editar_usuario(request, user_id):
    """Editar usuário existente"""
    # Verificar se o usuário é master (apenas master, não gerente)
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_master or user_profile.tipo_usuario != 'master':
            messages.error(request, 'Você não tem permissão para editar usuários.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    try:
        usuario = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Usuário não encontrado.')
        return redirect('gerenciar_usuarios')
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        password = request.POST.get('password', '').strip()
        tipo_usuario = request.POST.get('tipo_usuario', 'solicitante')
        is_active = request.POST.get('is_active') == 'on'
        
        # Validações
        if not username:
            messages.error(request, 'Username é obrigatório.')
            return render(request, 'fretes/editar_usuario.html', {'usuario': usuario})
        
        # Verificar se username já existe (exceto para o próprio usuário)
        if User.objects.filter(username=username).exclude(id=user_id).exists():
            messages.error(request, 'Username já existe.')
            return render(request, 'fretes/editar_usuario.html', {'usuario': usuario})
        
        # Verificar se email já existe (exceto para o próprio usuário)
        if email and User.objects.filter(email=email).exclude(id=user_id).exists():
            messages.error(request, 'Email já existe.')
            return render(request, 'fretes/editar_usuario.html', {'usuario': usuario})
        
        try:
            # Atualizar usuário
            usuario.username = username
            usuario.email = email
            usuario.first_name = first_name
            usuario.last_name = last_name
            usuario.is_active = is_active
            
            if password:
                usuario.set_password(password)
            
            usuario.save()
            
            # Atualizar perfil
            profile = usuario.userprofile
            profile.tipo_usuario = tipo_usuario
            if tipo_usuario == 'master':
                profile.is_master = True
                profile.tipo_acesso = 'completo'
            elif tipo_usuario == 'gerente':
                profile.is_master = False
                profile.tipo_acesso = 'completo'
            elif tipo_usuario == 'transportadora':
                profile.tipo_acesso = 'completo'
            else:  # solicitante
                profile.tipo_acesso = 'limitado'
            profile.save()
            
            messages.success(request, f'Usuário {username} atualizado com sucesso!')
            return redirect('gerenciar_usuarios')
            
        except Exception as e:
            messages.error(request, f'Erro ao atualizar usuário: {str(e)}')
            return render(request, 'fretes/editar_usuario.html', {'usuario': usuario})
    
    return render(request, 'fretes/editar_usuario.html', {'usuario': usuario})


@login_required(login_url='/login/')
def excluir_usuario(request, user_id):
    """Excluir usuário"""
    # Verificar se o usuário é master
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para excluir usuários.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    try:
        usuario = User.objects.get(id=user_id)
        
        # Não permitir excluir o próprio usuário
        if usuario == request.user:
            messages.error(request, 'Você não pode excluir seu próprio usuário.')
            return redirect('gerenciar_usuarios')
        
        username = usuario.username
        
        # Excluir usando SQL direto para evitar problemas com django_admin_log
        from django.db import connection
        
        try:
            with connection.cursor() as cursor:
                # Verificar se a tabela django_admin_log existe
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public' 
                        AND table_name = 'django_admin_log'
                    );
                """)
                
                table_exists = cursor.fetchone()[0]
                
                # Excluir logs do admin apenas se a tabela existir
                if table_exists:
                    cursor.execute("""
                        DELETE FROM django_admin_log 
                        WHERE user_id = %s
                    """, [usuario.id])
                
                # Excluir o usuário
                cursor.execute("""
                    DELETE FROM auth_user 
                    WHERE id = %s
                """, [usuario.id])
                
            messages.success(request, f'Usuário {username} excluído com sucesso!')
            
        except Exception as sql_error:
            # Se der erro no SQL, tentar método normal do Django
            try:
                usuario.delete()
                messages.success(request, f'Usuário {username} excluído com sucesso!')
            except Exception as django_error:
                # Se ambos falharem, mostrar erro
                messages.error(request, f'Erro ao excluir usuário: {str(django_error)}')
        
    except User.DoesNotExist:
        messages.error(request, 'Usuário não encontrado.')
    except Exception as e:
        messages.error(request, f'Erro ao excluir usuário: {str(e)}')
    
    return redirect('gerenciar_usuarios')


# Views para relatórios
@login_required(login_url='/login/')
def meus_fretes_relatorio_excel(request):
    """Gerar relatório Excel dos fretes do usuário"""
    status = request.GET.get('status')
    qs = FreteRequest.objects.select_related(
        'usuario', 'origem'
    ).prefetch_related('destinos').filter(usuario=request.user)
    
    if status in ['pendente', 'aguardando_aprovacao', 'aprovado', 'rejeitado', 'cotacao_enviada', 'cotacao_recebida', 'cotacao_aprovada', 'finalizado']:
        qs = qs.filter(status=status)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fretes'
    
    # Cabeçalho
    headers = [
        'Solicitante', 'Data Solicitação da Coleta', 'Origem Loja', 
        'Origem Endereço', 'Origem Cidade', 'Origem Estado', 'Origem CEP',
        'Destino Loja', 'Destino Endereço', 'Destino Cidade', 
        'Destino Estado', 'Destino CEP', 'Volume', 'Status'
    ]
    ws.append(headers)
    
    # Dados
    for frete in qs:
        for destino in frete.destinos.all():
            ws.append([
                frete.usuario.username,
                frete.data_criacao.strftime('%d/%m/%Y %H:%M'),
                frete.origem.nome if frete.origem else '',
                frete.origem.endereco if frete.origem else '',
                frete.origem.municipio if frete.origem else '',
                frete.origem.estado if frete.origem else '',
                frete.origem.cep if frete.origem else '',
                destino.loja or '',
                destino.endereco or '',
                destino.cidade or '',
                destino.estado or '',
                destino.cep or '',
                destino.volume,
                frete.status
            ])
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fretes_relatorio.xlsx'
    wb.save(response)
    return response


def health_check(request):
    """Endpoint para verificação de saúde do sistema"""
    return JsonResponse({
        'status': 'ok',
        'message': 'Portal de Fretes está funcionando',
        'timestamp': timezone.now().isoformat()
    })


@login_required(login_url='/login/')
@require_POST
def cancelar_frete(request, frete_id):
    """View para cancelar um frete com justificativa"""
    try:
        frete = get_object_or_404(FreteRequest, id=frete_id)
        
        # Verificar se o usuário pode cancelar este frete
        if frete.usuario != request.user and not request.user.userprofile.is_usuario_master:
            return JsonResponse({
                'success': False,
                'message': 'Você não tem permissão para cancelar este frete.'
            }, status=403)
        
        # Verificar se o frete pode ser cancelado
        if frete.status in ['finalizado', 'cancelado']:
            return JsonResponse({
                'success': False,
                'message': 'Este frete não pode ser cancelado.'
            }, status=400)
        
        # Obter a justificativa do POST
        motivo_cancelamento = request.POST.get('motivo_cancelamento', '').strip()
        
        if not motivo_cancelamento:
            return JsonResponse({
                'success': False,
                'message': 'É obrigatório informar o motivo do cancelamento.'
            }, status=400)
        
        # Cancelar o frete
        frete.status = 'cancelado'
        frete.motivo_cancelamento = motivo_cancelamento
        frete.usuario_cancelamento = request.user
        frete.data_cancelamento = timezone.now()
        frete.save()
        
        # Log do cancelamento
        print(f"LOG CANCELAMENTO - Frete #{frete.id} cancelado por {request.user.username} em {timezone.now()}")
        print(f"Motivo: {motivo_cancelamento}")
        
        return JsonResponse({
            'success': True,
            'message': 'Frete cancelado com sucesso.'
        })
        
    except Exception as e:
        print(f"Erro ao cancelar frete: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Erro interno do servidor. Tente novamente.'
        }, status=500)


# Form classes
class FreteForm(forms.ModelForm):
    """Form para criação de fretes"""
    origem = forms.ChoiceField(choices=[], label='Origem')
    destino = forms.MultipleChoiceField(choices=[], label='Destino')
    horario_coleta = forms.DateTimeField(
        label='Horário de Coleta',
        widget=forms.DateTimeInput(attrs={'type': 'datetime-local'})
    )

    class Meta:
        model = FreteRequest
        fields = ['descricao']


# Views para reset de senha
def forgot_password(request):
    """View para solicitar reset de senha"""
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = User.objects.get(email=email)
                # Gerar token e enviar email
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(force_bytes(user.pk))
                
                # URL para reset de senha
                reset_url = request.build_absolute_uri(
                    reverse('reset_password', kwargs={'uidb64': uid, 'token': token})
                )
                
                # Tentar enviar email, mas não falhar se não estiver configurado
                try:
                    # Enviar email
                    subject = 'Reset de Senha - Portal de Fretes'
                    message = f"""
                    Olá {user.username},
                    
                    Você solicitou um reset de senha para sua conta no Portal de Fretes.
                    
                    Clique no link abaixo para redefinir sua senha:
                    {reset_url}
                    
                    Este link é válido por 24 horas.
                    
                    Se você não solicitou este reset, ignore este email.
                    
                    Atenciosamente,
                    Equipe Portal de Fretes
                    """
                    
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [email],
                        fail_silently=True,  # Não falhar se email não estiver configurado
                    )
                    
                    messages.success(request, '✅ Email de reset enviado com sucesso! Verifique sua caixa de entrada e spam.')
                    
                except Exception as e:
                    # Se falhar o envio de email, mostrar o link diretamente
                    messages.success(request, f'✅ Link de reset gerado! Como o email não está configurado, copie o link abaixo: {reset_url}')
                
                return redirect('login')
                
            except User.DoesNotExist:
                messages.error(request, '❌ Email não encontrado em nosso sistema. Verifique se digitou corretamente.')
    else:
        form = PasswordResetForm()
    
    return render(request, 'fretes/forgot_password.html', {'form': form})


def reset_password(request, uidb64, token):
    """View para resetar senha com token"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, '✅ Senha redefinida com sucesso! Faça login com sua nova senha.')
                return redirect('login')
        else:
            form = SetPasswordForm(user)
        
        return render(request, 'fretes/reset_password.html', {'form': form})
    else:
        messages.error(request, '❌ Link inválido ou expirado. Solicite um novo reset de senha.')
        return redirect('forgot_password')

@login_required(login_url='/login/')
def fretes_aprovados_gerente(request):
    """Lista de fretes aprovados/rejeitados pelo gerente"""
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_gerente() and not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para acessar esta página.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar cotações aprovadas ou rejeitadas pelo usuário atual
    cotacoes = CotacaoFrete.objects.filter(
        aprovador=request.user,
        status__in=['aprovada', 'rejeitada']
    ).select_related('frete__usuario', 'frete__origem', 'transportadora').prefetch_related('frete__destinos').order_by('-data_aprovacao')
    
    # Estatísticas
    total_cotacoes = cotacoes.count()
    aprovadas = cotacoes.filter(status='aprovada').count()
    rejeitadas = cotacoes.filter(status='rejeitada').count()
    
    context = {
        'cotacoes': cotacoes,
        'total_cotacoes': total_cotacoes,
        'aprovadas': aprovadas,
        'rejeitadas': rejeitadas,
        'user_profile': user_profile,
    }
    
    return render(request, 'fretes/fretes_aprovados_gerente.html', context)


@login_required(login_url='/login/')
def fretes_aprovados_gerente_excel(request):
    """Exportar histórico de decisões do gerente para Excel"""
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_gerente() and not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para acessar esta página.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar cotações aprovadas ou rejeitadas pelo usuário atual
    cotacoes = CotacaoFrete.objects.filter(
        aprovador=request.user,
        status__in=['aprovada', 'rejeitada']
    ).select_related('frete__usuario', 'frete__origem', 'transportadora').prefetch_related('frete__destinos').order_by('-data_aprovacao')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico de Decisões'
    
    # Cabeçalho
    headers = [
        'ID Frete', 'Transportadora', 'Solicitante', 'Email Solicitante',
        'Origem Loja', 'Origem Endereço', 'Origem Cidade', 'Origem Estado', 'Origem CEP',
        'Destino Loja', 'Destino Endereço', 'Destino Cidade', 'Destino Estado', 'Destino CEP', 'Volume',
        'Valor Frete', 'Valor Pedágio', 'Valor Ajudante', 'Valor Total',
        'Status', 'Data Decisão', 'Observações Cotação', 'Observações Aprovação'
    ]
    ws.append(headers)
    
    # Dados
    for cotacao in cotacoes:
        for destino in cotacao.frete.destinos.all():
            ws.append([
                cotacao.frete.id,
                cotacao.transportadora.nome,
                cotacao.frete.usuario.username,
                cotacao.frete.usuario.email,
                cotacao.frete.origem.nome if cotacao.frete.origem else '',
                cotacao.frete.origem.endereco if cotacao.frete.origem else '',
                cotacao.frete.origem.municipio if cotacao.frete.origem else '',
                cotacao.frete.origem.estado if cotacao.frete.origem else '',
                cotacao.frete.origem.cep if cotacao.frete.origem else '',
                destino.loja or '',
                destino.endereco or '',
                destino.cidade or '',
                destino.estado or '',
                destino.cep or '',
                destino.volume,
                cotacao.valor_frete or 0,
                cotacao.valor_pedagio or 0,
                cotacao.valor_ajudante or 0,
                cotacao.valor_total or 0,
                'Aprovada' if cotacao.status == 'aprovada' else 'Rejeitada',
                cotacao.data_aprovacao.strftime('%d/%m/%Y %H:%M') if cotacao.data_aprovacao else '',
                cotacao.observacoes_cotacao or '',
                cotacao.observacoes_aprovacao or ''
            ])
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historico_decisoes_gerente.xlsx'
    wb.save(response)
    return response


@login_required(login_url='/login/')
def cotacoes_historico_transportadora(request):
    """Lista de todas as cotações da transportadora (todos os status)"""
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_transportadora() and not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para acessar esta página.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar todas as cotações da transportadora (todos os status)
    if user_profile.is_usuario_master():
        # Master vê todas as cotações
        cotacoes = CotacaoFrete.objects.all().select_related(
            'frete__usuario', 'frete__origem', 'transportadora', 'aprovador'
        ).prefetch_related('frete__destinos').order_by('-data_cotacao', '-data_aprovacao')
    else:
        # Transportadora vê apenas suas cotações
        cotacoes = CotacaoFrete.objects.filter(
            transportadora=user_profile.transportadora
        ).select_related('frete__usuario', 'frete__origem', 'transportadora', 'aprovador').prefetch_related('frete__destinos').order_by('-data_cotacao', '-data_aprovacao')
    
    # Estatísticas
    total_cotacoes = cotacoes.count()
    pendentes = cotacoes.filter(status='pendente').count()
    enviadas = cotacoes.filter(status='enviada').count()
    aprovadas = cotacoes.filter(status='aprovada').count()
    rejeitadas = cotacoes.filter(status='rejeitada').count()
    rejeitadas_transportadora = cotacoes.filter(status='rejeitada_transportadora').count()
    
    # Valor total das cotações aprovadas
    valor_total_aprovadas = cotacoes.filter(status='aprovada').aggregate(
        total=models.Sum('valor_total')
    )['total'] or 0
    
    context = {
        'cotacoes': cotacoes,
        'total_cotacoes': total_cotacoes,
        'pendentes': pendentes,
        'enviadas': enviadas,
        'aprovadas': aprovadas,
        'rejeitadas': rejeitadas,
        'rejeitadas_transportadora': rejeitadas_transportadora,
        'valor_total_aprovadas': valor_total_aprovadas,
        'user_profile': user_profile,
    }
    
    return render(request, 'fretes/cotacoes_historico_transportadora.html', context)


@login_required(login_url='/login/')
def cotacoes_historico_transportadora_excel(request):
    """Exportar histórico de cotações da transportadora para Excel"""
    try:
        user_profile = request.user.userprofile
        if not user_profile.is_transportadora() and not user_profile.is_usuario_master():
            messages.error(request, 'Você não tem permissão para acessar esta página.')
            return redirect('home')
    except UserProfile.DoesNotExist:
        messages.error(request, 'Perfil de usuário não encontrado.')
        return redirect('home')
    
    # Buscar todas as cotações da transportadora (todos os status)
    if user_profile.is_usuario_master():
        # Master vê todas as cotações
        cotacoes = CotacaoFrete.objects.all().select_related(
            'frete__usuario', 'frete__origem', 'transportadora', 'aprovador'
        ).prefetch_related('frete__destinos').order_by('-data_cotacao', '-data_aprovacao')
    else:
        # Transportadora vê apenas suas cotações
        cotacoes = CotacaoFrete.objects.filter(
            transportadora=user_profile.transportadora
        ).select_related('frete__usuario', 'frete__origem', 'transportadora', 'aprovador').prefetch_related('frete__destinos').order_by('-data_cotacao', '-data_aprovacao')
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Histórico de Cotações'
    
    # Cabeçalho
    headers = [
        'ID Frete', 'Transportadora', 'Solicitante', 'Email Solicitante',
        'Origem Loja', 'Origem Endereço', 'Origem Cidade', 'Origem Estado', 'Origem CEP',
        'Destino Loja', 'Destino Endereço', 'Destino Cidade', 'Destino Estado', 'Destino CEP', 'Volume',
        'Valor Frete', 'Valor Pedágio', 'Valor Ajudante', 'Valor Total',
        'Status', 'Aprovador', 'Data Aprovação', 'Observações Cotação', 'Observações Aprovação'
    ]
    ws.append(headers)
    
    # Dados
    for cotacao in cotacoes:
        for destino in cotacao.frete.destinos.all():
            ws.append([
                cotacao.frete.id,
                cotacao.transportadora.nome,
                cotacao.frete.usuario.username,
                cotacao.frete.usuario.email,
                cotacao.frete.origem.nome if cotacao.frete.origem else '',
                cotacao.frete.origem.endereco if cotacao.frete.origem else '',
                cotacao.frete.origem.municipio if cotacao.frete.origem else '',
                cotacao.frete.origem.estado if cotacao.frete.origem else '',
                cotacao.frete.origem.cep if cotacao.frete.origem else '',
                destino.loja or '',
                destino.endereco or '',
                destino.cidade or '',
                destino.estado or '',
                destino.cep or '',
                destino.volume,
                cotacao.valor_frete or 0,
                cotacao.valor_pedagio or 0,
                cotacao.valor_ajudante or 0,
                cotacao.valor_total or 0,
                'Aprovada' if cotacao.status == 'aprovada' else 'Rejeitada',
                cotacao.aprovador.username if cotacao.aprovador else '',
                cotacao.data_aprovacao.strftime('%d/%m/%Y %H:%M') if cotacao.data_aprovacao else '',
                cotacao.observacoes_cotacao or '',
                cotacao.observacoes_aprovacao or ''
            ])
    
    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historico_cotacoes_transportadora.xlsx'
    wb.save(response)
    return response


@login_required(login_url='/login/')
def gerenciar_transportadoras(request):
    """Página para gerenciar transportadoras - apenas para usuários master"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            messages.error(request, 'Apenas usuários master podem gerenciar transportadoras.')
            return redirect('home')
        
        # OTIMIZADO: Cache das transportadoras por 30 minutos
        from django.core.cache import cache
        cache_key = 'transportadoras_list'
        transportadoras = cache.get(cache_key)
        
        if transportadoras is None:
            transportadoras = Transportadora.objects.all().order_by('nome')
            cache.set(cache_key, transportadoras, 1800)  # 30 minutos
        
        return render(request, 'fretes/gerenciar_transportadoras.html', {
            'transportadoras': transportadoras
        })
        
    except Exception as e:
        messages.error(request, f'Erro ao carregar transportadoras: {str(e)}')
        return redirect('home')


@login_required(login_url='/login/')
@require_POST
def criar_transportadora_ajax(request):
    """View AJAX para criar nova transportadora - apenas para usuários master"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            return JsonResponse({
                'success': False,
                'error': 'Apenas usuários master podem criar transportadoras.'
            })
        
        # Obter dados do formulário
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip()
        
        # Validações
        if not nome:
            return JsonResponse({
                'success': False,
                'error': 'Nome da transportadora é obrigatório.'
            })
        
        if not email:
            return JsonResponse({
                'success': False,
                'error': 'Email é obrigatório.'
            })
        
        # Verificar se já existe transportadora com esse nome
        if Transportadora.objects.filter(nome__iexact=nome).exists():
            return JsonResponse({
                'success': False,
                'error': f'Já existe uma transportadora com o nome "{nome}".'
            })
        
        # Verificar se já existe transportadora com esse email
        if Transportadora.objects.filter(email__iexact=email).exists():
            return JsonResponse({
                'success': False,
                'error': f'Já existe uma transportadora com o email "{email}".'
            })
        
        # Criar transportadora
        transportadora = Transportadora.objects.create(
            nome=nome,
            email=email
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Transportadora "{transportadora.nome}" criada com sucesso!',
            'transportadora_id': transportadora.id,
            'transportadora_nome': transportadora.nome
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        })


@login_required(login_url='/login/')
@require_POST
def editar_transportadora_ajax(request):
    """View AJAX para editar transportadora - apenas para usuários master"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            return JsonResponse({
                'success': False,
                'error': 'Apenas usuários master podem editar transportadoras.'
            })
        
        # Obter dados do formulário
        transportadora_id = request.POST.get('id')
        nome = request.POST.get('nome', '').strip()
        email = request.POST.get('email', '').strip()
        
        # Validações
        if not transportadora_id:
            return JsonResponse({
                'success': False,
                'error': 'ID da transportadora é obrigatório.'
            })
        
        if not nome:
            return JsonResponse({
                'success': False,
                'error': 'Nome da transportadora é obrigatório.'
            })
        
        if not email:
            return JsonResponse({
                'success': False,
                'error': 'Email é obrigatório.'
            })
        
        # Buscar transportadora
        try:
            transportadora = Transportadora.objects.get(id=transportadora_id)
        except Transportadora.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Transportadora não encontrada.'
            })
        
        # Verificar se já existe transportadora com esse nome (excluindo a atual)
        if Transportadora.objects.filter(nome__iexact=nome).exclude(id=transportadora_id).exists():
            return JsonResponse({
                'success': False,
                'error': f'Já existe uma transportadora com o nome "{nome}".'
            })
        
        # Verificar se já existe transportadora com esse email (excluindo a atual)
        if Transportadora.objects.filter(email__iexact=email).exclude(id=transportadora_id).exists():
            return JsonResponse({
                'success': False,
                'error': f'Já existe uma transportadora com o email "{email}".'
            })
        
        # Atualizar transportadora
        transportadora.nome = nome
        transportadora.email = email
        transportadora.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Transportadora "{transportadora.nome}" atualizada com sucesso!',
            'transportadora_id': transportadora.id,
            'transportadora_nome': transportadora.nome
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        })


@login_required(login_url='/login/')
@require_POST
def excluir_transportadora_ajax(request):
    """View AJAX para excluir transportadora - apenas para usuários master"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            return JsonResponse({
                'success': False,
                'error': 'Apenas usuários master podem excluir transportadoras.'
            })
        
        # Obter ID da transportadora
        transportadora_id = request.POST.get('id')
        
        if not transportadora_id:
            return JsonResponse({
                'success': False,
                'error': 'ID da transportadora é obrigatório.'
            })
        
        # Buscar transportadora
        try:
            transportadora = Transportadora.objects.get(id=transportadora_id)
        except Transportadora.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Transportadora não encontrada.'
            })
        
        # Verificar se há cotações associadas
        from fretes.models import CotacaoFrete
        cotacoes_count = CotacaoFrete.objects.filter(transportadora=transportadora).count()
        
        if cotacoes_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Não é possível excluir a transportadora "{transportadora.nome}" pois ela possui {cotacoes_count} cotação(ões) associada(s).'
            })
        
        # Excluir transportadora
        nome_transportadora = transportadora.nome
        transportadora.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Transportadora "{nome_transportadora}" excluída com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        })


@login_required(login_url='/login/')
def gerenciar_usuarios_transportadora(request, transportadora_id):
    """Página para gerenciar usuários de uma transportadora específica"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            messages.error(request, 'Apenas usuários master podem gerenciar usuários de transportadoras.')
            return redirect('home')
        
        # Buscar transportadora
        try:
            transportadora = Transportadora.objects.get(id=transportadora_id)
        except Transportadora.DoesNotExist:
            messages.error(request, 'Transportadora não encontrada.')
            return redirect('gerenciar_transportadoras')
        
        # Buscar usuários associados à transportadora
        usuarios_associados = UserProfile.objects.filter(
            transportadora=transportadora,
            tipo_usuario='transportadora'
        ).select_related('user')
        
        # Buscar usuários disponíveis para associar
        # Incluir usuários que são solicitantes, gerentes ou transportadoras sem transportadora associada
        usuarios_disponiveis = User.objects.filter(
            userprofile__tipo_usuario__in=['solicitante', 'gerente', 'transportadora']
        ).exclude(
            userprofile__transportadora__isnull=False
        ).select_related('userprofile')
        
        return render(request, 'fretes/gerenciar_usuarios_transportadora.html', {
            'transportadora': transportadora,
            'usuarios_associados': usuarios_associados,
            'usuarios_disponiveis': usuarios_disponiveis
        })
        
    except Exception as e:
        messages.error(request, f'Erro ao carregar usuários: {str(e)}')
        return redirect('gerenciar_transportadoras')


@login_required(login_url='/login/')
@require_POST
def associar_usuario_transportadora_ajax(request):
    """View AJAX para associar usuário à transportadora"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            return JsonResponse({
                'success': False,
                'error': 'Apenas usuários master podem associar usuários.'
            })
        
        # Obter dados
        user_id = request.POST.get('user_id')
        transportadora_id = request.POST.get('transportadora_id')
        
        if not user_id or not transportadora_id:
            return JsonResponse({
                'success': False,
                'error': 'Dados incompletos.'
            })
        
        # Buscar usuário e transportadora
        try:
            user = User.objects.get(id=user_id)
            transportadora = Transportadora.objects.get(id=transportadora_id)
        except (User.DoesNotExist, Transportadora.DoesNotExist):
            return JsonResponse({
                'success': False,
                'error': 'Usuário ou transportadora não encontrados.'
            })
        
        # Atualizar perfil do usuário
        profile, created = UserProfile.objects.get_or_create(
            user=user,
            defaults={
                'tipo_usuario': 'transportadora',
                'tipo_acesso': 'completo',
                'transportadora': transportadora
            }
        )
        
        # Sempre atualizar para garantir que está correto
        profile.tipo_usuario = 'transportadora'
        profile.tipo_acesso = 'completo'
        profile.transportadora = transportadora
        profile.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Usuário {user.username} associado à transportadora {transportadora.nome} com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        })


@login_required(login_url='/login/')
@require_POST
def desassociar_usuario_transportadora_ajax(request):
    """View AJAX para desassociar usuário da transportadora"""
    try:
        # Verificar se o usuário é master (apenas master, não gerente)
        if not hasattr(request.user, 'userprofile') or not request.user.userprofile.is_master or request.user.userprofile.tipo_usuario != 'master':
            return JsonResponse({
                'success': False,
                'error': 'Apenas usuários master podem desassociar usuários.'
            })
        
        # Obter dados
        user_id = request.POST.get('user_id')
        
        if not user_id:
            return JsonResponse({
                'success': False,
                'error': 'ID do usuário é obrigatório.'
            })
        
        # Buscar usuário
        try:
            user = User.objects.get(id=user_id)
            profile = user.userprofile
        except User.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Usuário não encontrado.'
            })
        
        # Desassociar da transportadora
        transportadora_nome = profile.transportadora.nome if profile.transportadora else 'N/A'
        profile.transportadora = None
        profile.tipo_usuario = 'solicitante'
        profile.tipo_acesso = 'limitado'
        profile.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Usuário {user.username} desassociado da transportadora {transportadora_nome} com sucesso!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        })