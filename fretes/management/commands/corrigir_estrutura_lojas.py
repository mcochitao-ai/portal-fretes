from django.core.management.base import BaseCommand
from fretes.models import Loja
import os

class Command(BaseCommand):
    help = 'Corrige a estrutura das lojas baseado na estrutura real do Excel'

    def handle(self, *args, **options):
        self.stdout.write('🔍 ANALISANDO ESTRUTURA REAL DO EXCEL...')
        self.stdout.write('=' * 60)
        
        # Verificar se o arquivo existe
        file_path = 'Lojas Portal.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'❌ Arquivo {file_path} não encontrado!'))
            return
        
        try:
            from openpyxl import load_workbook
            
            # Carregar o arquivo Excel
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # Mostrar cabeçalho
            self.stdout.write(f'📋 CABEÇALHO (linha 1):')
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            for i, header in enumerate(header_row):
                self.stdout.write(f'   Coluna {i}: "{header}"')
            
            # Mostrar algumas linhas de dados para entender a estrutura
            self.stdout.write(f'\n📋 PRIMEIRAS 3 LINHAS DE DADOS:')
            for row_num in range(2, 5):
                row_data = list(worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                self.stdout.write(f'\n   Linha {row_num}:')
                for i, cell_value in enumerate(row_data):
                    if cell_value is not None and str(cell_value).strip():
                        self.stdout.write(f'     Coluna {i}: "{cell_value}"')
            
            # Limpar lojas existentes
            self.stdout.write(f'\n🗑️ Removendo lojas existentes...')
            Loja.objects.all().delete()
            
            # Importar com estrutura correta baseada no que vimos
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            importadas = 0
            
            for i, row in enumerate(rows, start=2):
                if not row[0]:  # Se a primeira coluna estiver vazia, pular
                    continue
                    
                try:
                    # Estrutura correta baseada no exemplo:
                    # Coluna 0: Número da loja (18) -> Nome: "Loja 18"
                    # Coluna 1: Endereço (RUA SETE DE SETEMBRO) -> Endereço + número
                    # Coluna 2: Município (mas está vindo como "-") -> Procurar em outras colunas
                    # Coluna 3: Estado (BLUMENAU) -> Converter para sigla (SC)
                    # Coluna 4: CEP (mas está vindo como SANTA CATARINA) -> Procurar CEP real
                    
                    # Nome da loja
                    numero_loja = str(row[0]) if row[0] else ''
                    nome = f"Loja {numero_loja}" if numero_loja else ''
                    
                    # Endereço completo
                    endereco_base = str(row[1]) if row[1] else ''
                    endereco = endereco_base
                    
                    # Procurar número do endereço e município nas colunas seguintes
                    municipio = ''
                    estado = ''
                    cep = ''
                    regional = ''
                    numero_endereco = ''
                    
                    # Mapear estados para siglas
                    estado_map = {
                        'SÃO PAULO': 'SP',
                        'RIO DE JANEIRO': 'RJ', 
                        'MINAS GERAIS': 'MG',
                        'BAHIA': 'BA',
                        'PARANÁ': 'PR',
                        'RIO GRANDE DO SUL': 'RS',
                        'PERNAMBUCO': 'PE',
                        'CEARÁ': 'CE',
                        'PARÁ': 'PA',
                        'SANTA CATARINA': 'SC',
                        'GOIÁS': 'GO',
                        'MARANHÃO': 'MA',
                        'ESPÍRITO SANTO': 'ES',
                        'PIAUÍ': 'PI',
                        'ALAGOAS': 'AL',
                        'TOCANTINS': 'TO',
                        'RIO GRANDE DO NORTE': 'RN',
                        'ACRE': 'AC',
                        'AMAPÁ': 'AP',
                        'AMAZONAS': 'AM',
                        'MATO GROSSO': 'MT',
                        'MATO GROSSO DO SUL': 'MS',
                        'RONDÔNIA': 'RO',
                        'RORAIMA': 'RR',
                        'SERGIPE': 'SE',
                        'DISTRITO FEDERAL': 'DF',
                        'BLUMENAU': 'SC'  # Blumenau está em SC
                    }
                    
                    # Analisar cada coluna
                    for col_idx in range(2, len(row)):
                        if row[col_idx] is not None:
                            valor = str(row[col_idx]).strip()
                            if valor and valor != '-':
                                # Se contém apenas números, pode ser número do endereço
                                if valor.isdigit() and len(valor) <= 5:
                                    numero_endereco = valor
                                # Se contém números e hífen, pode ser CEP
                                elif any(c.isdigit() for c in valor) and ('-' in valor or len(valor) >= 8):
                                    cep = valor[:15]
                                # Se é um estado conhecido, converter para sigla
                                elif valor.upper() in estado_map:
                                    estado = estado_map[valor.upper()]
                                # Se contém "REGIONAL" ou "YCR", é regional
                                elif 'REGIONAL' in valor.upper() or 'YCR' in valor.upper():
                                    regional = valor[:100]
                                # Se não é estado nem regional, pode ser município
                                elif not municipio and len(valor) > 2:
                                    municipio = valor[:100]
                    
                    # Se encontrou número do endereço, adicionar ao endereço
                    if numero_endereco and endereco:
                        endereco = f"{endereco}, {numero_endereco}"
                    
                    # Se não encontrou município, tentar extrair do estado
                    if not municipio and estado:
                        if estado == 'SC' and 'BLUMENAU' in str(row[3]).upper():
                            municipio = 'Blumenau'
                        # Adicionar mais mapeamentos conforme necessário
                    
                    # Procurar coordenadas
                    latitude = None
                    longitude = None
                    for col_idx in range(2, len(row)):
                        if row[col_idx] is not None:
                            try:
                                valor = float(row[col_idx])
                                if -90 <= valor <= 90:  # Latitude válida
                                    if latitude is None:
                                        latitude = valor
                                elif -180 <= valor <= 180:  # Longitude válida
                                    if longitude is None:
                                        longitude = valor
                            except (ValueError, TypeError):
                                continue
                    
                    if not nome:
                        continue
                    
                    # Criar loja
                    loja = Loja.objects.create(
                        nome=nome,
                        endereco=endereco,
                        numero='',
                        municipio=municipio,
                        estado=estado,
                        cep=cep,
                        regional=regional,
                        latitude=latitude,
                        longitude=longitude
                    )
                    
                    importadas += 1
                    if importadas <= 5:  # Mostrar apenas as primeiras 5
                        self.stdout.write(f'✅ Loja {loja.nome}: {loja.endereco}, {loja.municipio}/{loja.estado}')
                
                except Exception as e:
                    self.stdout.write(f'❌ Erro na linha {i}: {e}')
            
            self.stdout.write(f'\n📈 RESUMO:')
            self.stdout.write(f'   - Lojas importadas: {importadas}')
            self.stdout.write(f'   - Total no banco: {Loja.objects.count()}')
            
            if importadas > 0:
                self.stdout.write(self.style.SUCCESS('🎉 Importação concluída!'))
            else:
                self.stdout.write(self.style.WARNING('⚠️ Nenhuma loja foi importada'))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro: {e}'))
            import traceback
            self.stdout.write(traceback.format_exc())
        
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('🏁 ANÁLISE E CORREÇÃO COMPLETA!')
