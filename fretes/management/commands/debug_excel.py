from django.core.management.base import BaseCommand
import os

class Command(BaseCommand):
    help = 'Debug da estrutura do arquivo Excel'

    def handle(self, *args, **options):
        self.stdout.write('🔍 DEBUG DA ESTRUTURA DO EXCEL...')
        self.stdout.write('=' * 60)
        
        # Verificar se o arquivo existe
        file_path = 'Lojas Portal.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'❌ Arquivo {file_path} não encontrado!'))
            return
        
        try:
            from openpyxl import load_workbook
            
            # Carregar o arquivo Excel
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            self.stdout.write(f'📊 Arquivo Excel carregado')
            self.stdout.write(f'📋 Nome da planilha: {worksheet.title}')
            self.stdout.write(f'📏 Dimensões: {worksheet.max_row} linhas x {worksheet.max_column} colunas')
            
            # Mostrar cabeçalho (primeira linha)
            self.stdout.write(f'\n📋 CABEÇALHO (linha 1):')
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            for i, header in enumerate(header_row):
                self.stdout.write(f'   Coluna {i}: "{header}"')
            
            # Mostrar algumas linhas de dados
            self.stdout.write(f'\n📋 PRIMEIRAS 5 LINHAS DE DADOS:')
            for row_num in range(2, min(7, worksheet.max_row + 1)):
                row_data = list(worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                self.stdout.write(f'\n   Linha {row_num}:')
                for i, cell_value in enumerate(row_data):
                    if cell_value is not None:
                        self.stdout.write(f'     Coluna {i}: "{cell_value}" (tipo: {type(cell_value).__name__})')
            
            # Procurar por coordenadas
            self.stdout.write(f'\n🔍 PROCURANDO COORDENADAS...')
            for row_num in range(2, min(10, worksheet.max_row + 1)):
                row_data = list(worksheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0]
                self.stdout.write(f'\n   Linha {row_num}:')
                for i, cell_value in enumerate(row_data):
                    if cell_value is not None:
                        try:
                            valor = float(cell_value)
                            if -90 <= valor <= 90:  # Possível latitude
                                self.stdout.write(f'     Coluna {i}: {valor} (possível LATITUDE)')
                            elif -180 <= valor <= 180:  # Possível longitude
                                self.stdout.write(f'     Coluna {i}: {valor} (possível LONGITUDE)')
                        except (ValueError, TypeError):
                            pass
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro: {e}'))
            import traceback
            self.stdout.write(traceback.format_exc())
        
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('🏁 DEBUG COMPLETO!')
