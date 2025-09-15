from django.core.management.base import BaseCommand
from fretes.models import Loja
import os

class Command(BaseCommand):
    help = 'Importa lojas do arquivo Excel sem dependência do pandas'

    def handle(self, *args, **options):
        self.stdout.write('🔄 IMPORTANDO LOJAS DO EXCEL (versão simples)...')
        
        # Verificar se o arquivo existe
        file_path = 'Lojas Portal.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'❌ Arquivo {file_path} não encontrado!'))
            return
        
        try:
            # Usar openpyxl diretamente (já está no requirements.txt)
            from openpyxl import load_workbook
            
            # Carregar o arquivo Excel
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # Pular o cabeçalho (primeira linha)
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            self.stdout.write(f'📊 Arquivo Excel carregado: {len(rows)} lojas encontradas')
            
            # Contar lojas importadas
            importadas = 0
            existentes = 0
            
            for row in rows:
                if not row[0]:  # Se a primeira coluna estiver vazia, pular
                    continue
                    
                # Extrair dados da linha
                nome = str(row[0]) if row[0] else ''
                endereco = str(row[1]) if row[1] else ''
                numero = str(row[2]) if row[2] else ''
                municipio = str(row[3]) if row[3] else ''
                estado = str(row[4]) if row[4] else ''
                cep = str(row[5]) if row[5] else ''
                regional = str(row[6]) if row[6] else ''
                latitude = float(row[7]) if row[7] and row[7] != 'None' else None
                longitude = float(row[8]) if row[8] and row[8] != 'None' else None
                
                if not nome:  # Pular se não tiver nome
                    continue
                
                loja, created = Loja.objects.get_or_create(
                    nome=nome,
                    defaults={
                        'endereco': endereco,
                        'numero': numero,
                        'municipio': municipio,
                        'estado': estado,
                        'cep': cep,
                        'regional': regional,
                        'latitude': latitude,
                        'longitude': longitude
                    }
                )
                
                if created:
                    importadas += 1
                    self.stdout.write(f'✅ Loja {loja.nome} importada')
                else:
                    existentes += 1
            
            self.stdout.write(f'\n📈 RESUMO DA IMPORTAÇÃO:')
            self.stdout.write(f'   - Lojas importadas: {importadas}')
            self.stdout.write(f'   - Lojas já existentes: {existentes}')
            self.stdout.write(f'   - Total no banco: {Loja.objects.count()}')
            
            if importadas > 0:
                self.stdout.write(self.style.SUCCESS('🎉 Importação concluída com sucesso!'))
            else:
                self.stdout.write(self.style.WARNING('⚠️ Nenhuma loja nova foi importada (todas já existiam)'))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro na importação: {e}'))
            import traceback
            self.stdout.write(traceback.format_exc())
