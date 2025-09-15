from django.core.management.base import BaseCommand
from django.db import connection
from django.contrib.auth.models import User
from fretes.models import Loja, Transportadora, UserProfile
import os

class Command(BaseCommand):
    help = 'Setup completo do sistema - cria tudo de uma vez'

    def handle(self, *args, **options):
        self.stdout.write('🚀 SETUP COMPLETO DO SISTEMA')
        self.stdout.write('=' * 60)
        
        try:
            # 1. CRIAR ESTRUTURA DO BANCO
            self.stdout.write('🔧 Criando estrutura do banco...')
            self.criar_estrutura_banco()
            
            # 2. CRIAR USUÁRIO ADMIN
            self.stdout.write('👤 Criando usuário admin...')
            self.criar_admin()
            
            # 3. IMPORTAR LOJAS
            self.stdout.write('🏪 Importando lojas...')
            self.importar_lojas()
            
            # 4. CADASTRAR TRANSPORTADORAS
            self.stdout.write('🚛 Cadastrando transportadoras...')
            self.cadastrar_transportadoras()
            
            self.stdout.write(self.style.SUCCESS('🎉 SETUP COMPLETO COM SUCESSO!'))
            
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro no setup: {e}'))
            import traceback
            self.stdout.write(traceback.format_exc())
    
    def criar_estrutura_banco(self):
        """Cria a estrutura do banco de dados usando migrações do Django"""
        try:
            from django.core.management import call_command
            
            # Usar migrações do Django em vez de SQL manual
            call_command('migrate', verbosity=0, interactive=False)
            self.stdout.write('✅ Estrutura do banco criada via migrações')
            
        except Exception as e:
            self.stdout.write(f'⚠️ Erro ao executar migrações: {e}')
            # Fallback para SQL manual se necessário
            self.criar_estrutura_banco_manual()
    
    def criar_estrutura_banco_manual(self):
        """Fallback: Cria a estrutura do banco de dados manualmente"""
        try:
            with connection.cursor() as cursor:
                # Verificar se as tabelas já existem
                cursor.execute("""
                    SELECT table_name FROM information_schema.tables 
                    WHERE table_schema = 'public' AND table_name = 'auth_user'
                """)
                
                if cursor.fetchone():
                    self.stdout.write('✅ Tabelas já existem, pulando criação')
                    return
                
                # Criar tabelas apenas se não existirem
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS auth_user (
                        id SERIAL PRIMARY KEY,
                        password VARCHAR(128) NOT NULL,
                        last_login TIMESTAMP WITH TIME ZONE,
                        is_superuser BOOLEAN NOT NULL DEFAULT FALSE,
                        username VARCHAR(150) NOT NULL UNIQUE,
                        first_name VARCHAR(150) NOT NULL DEFAULT '',
                        last_name VARCHAR(150) NOT NULL DEFAULT '',
                        email VARCHAR(254) NOT NULL DEFAULT '',
                        is_staff BOOLEAN NOT NULL DEFAULT FALSE,
                        is_active BOOLEAN NOT NULL DEFAULT TRUE,
                        date_joined TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_transportadora (
                        id SERIAL PRIMARY KEY,
                        nome VARCHAR(100) NOT NULL,
                        email VARCHAR(254) NOT NULL
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_loja (
                        id SERIAL PRIMARY KEY,
                        nome VARCHAR(255) NOT NULL,
                        endereco TEXT NOT NULL,
                        numero VARCHAR(20) DEFAULT '',
                        municipio VARCHAR(100) DEFAULT '',
                        estado VARCHAR(10) DEFAULT '',
                        cep VARCHAR(15) DEFAULT '',
                        regional VARCHAR(100) DEFAULT '',
                        latitude DOUBLE PRECISION,
                        longitude DOUBLE PRECISION
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_freterequest (
                        id SERIAL PRIMARY KEY,
                        data_criacao TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW(),
                        descricao TEXT NOT NULL DEFAULT '',
                        usuario_id INTEGER NOT NULL REFERENCES auth_user(id) ON DELETE CASCADE,
                        transportadora_selecionada_id INTEGER REFERENCES fretes_transportadora(id) ON DELETE SET NULL,
                        origem_id INTEGER REFERENCES fretes_loja(id) ON DELETE SET NULL,
                        horario_coleta TIMESTAMP WITH TIME ZONE,
                        observacoes_origem TEXT DEFAULT '',
                        anexo_origem VARCHAR(100) DEFAULT '',
                        anexo_nota_fiscal VARCHAR(100) DEFAULT '',
                        tipo_veiculo VARCHAR(50) DEFAULT '',
                        precisa_ajudante BOOLEAN DEFAULT FALSE,
                        quantidade_ajudantes INTEGER DEFAULT 0,
                        nota_fiscal_emitida BOOLEAN DEFAULT FALSE,
                        quem_paga_frete VARCHAR(100) DEFAULT '',
                        status VARCHAR(50) DEFAULT 'pendente',
                        centro_custo VARCHAR(100) DEFAULT '',
                        aprovador_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                        data_aprovacao TIMESTAMP WITH TIME ZONE,
                        observacoes_aprovacao TEXT DEFAULT '',
                        justificativa_rejeicao TEXT DEFAULT '',
                        motivo_cancelamento TEXT DEFAULT '',
                        usuario_cancelamento_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                        data_cancelamento TIMESTAMP WITH TIME ZONE,
                        transportadora_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                        valor_frete DECIMAL(10,2),
                        valor_pedagio DECIMAL(10,2),
                        valor_ajudante DECIMAL(10,2),
                        valor_total DECIMAL(10,2),
                        data_cotacao TIMESTAMP WITH TIME ZONE,
                        observacoes_cotacao TEXT DEFAULT '',
                        motivo_rejeicao_transportadora TEXT DEFAULT ''
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_destino (
                        id SERIAL PRIMARY KEY,
                        endereco VARCHAR(255) NOT NULL,
                        cidade VARCHAR(100) NOT NULL,
                        estado VARCHAR(2) NOT NULL,
                        cep VARCHAR(10) NOT NULL,
                        volume INTEGER DEFAULT 1,
                        frete_id INTEGER NOT NULL REFERENCES fretes_freterequest(id) ON DELETE CASCADE,
                        loja VARCHAR(100) DEFAULT '',
                        numero VARCHAR(20) DEFAULT '',
                        observacao TEXT DEFAULT '',
                        anexo_destino VARCHAR(100) DEFAULT '',
                        data_entrega TIMESTAMP WITH TIME ZONE
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_cotacaofrete (
                        id SERIAL PRIMARY KEY,
                        frete_id INTEGER NOT NULL REFERENCES fretes_freterequest(id) ON DELETE CASCADE,
                        transportadora_id INTEGER NOT NULL REFERENCES fretes_transportadora(id) ON DELETE CASCADE,
                        valor_frete DECIMAL(10,2),
                        valor_pedagio DECIMAL(10,2),
                        valor_ajudante DECIMAL(10,2),
                        valor_total DECIMAL(10,2),
                        status VARCHAR(30) DEFAULT 'pendente',
                        data_cotacao TIMESTAMP WITH TIME ZONE,
                        observacoes_cotacao TEXT,
                        motivo_rejeicao_transportadora TEXT,
                        aprovador_id INTEGER REFERENCES auth_user(id) ON DELETE SET NULL,
                        data_aprovacao TIMESTAMP WITH TIME ZONE,
                        observacoes_aprovacao TEXT,
                        justificativa_rejeicao TEXT
                    );
                """)
                
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS fretes_userprofile (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL REFERENCES auth_user(id) ON DELETE CASCADE,
                        tipo_acesso VARCHAR(20) DEFAULT 'limitado',
                        is_master BOOLEAN DEFAULT FALSE,
                        tipo_usuario VARCHAR(20) DEFAULT 'solicitante',
                        transportadora_id INTEGER REFERENCES fretes_transportadora(id) ON DELETE SET NULL
                    );
                """)
                
                # Criar índices apenas se não existirem
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS fretes_freterequest_usuario_id_idx 
                    ON fretes_freterequest (usuario_id);
                """)
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS fretes_freterequest_transportadora_selecionada_id_idx 
                    ON fretes_freterequest (transportadora_selecionada_id);
                """)
                cursor.execute("""
                    CREATE INDEX IF NOT EXISTS fretes_destino_frete_id_idx 
                    ON fretes_destino (frete_id);
                """)
                cursor.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS fretes_userprofile_user_id_key 
                    ON fretes_userprofile (user_id);
                """)
                
                self.stdout.write('✅ Estrutura do banco criada manualmente')
                
        except Exception as e:
            self.stdout.write(f'❌ Erro ao criar estrutura do banco: {e}')
            raise
    
    def criar_admin(self):
        """Cria usuário admin"""
        admin_user, created = User.objects.get_or_create(
            username='admin',
            defaults={
                'email': 'admin@portal.com',
                'is_staff': True,
                'is_superuser': True,
                'is_active': True
            }
        )
        
        # Sempre atualizar senha e permissões
        admin_user.set_password('admin123')
        admin_user.is_staff = True
        admin_user.is_superuser = True
        admin_user.is_active = True
        admin_user.save()
        
        if created:
            self.stdout.write('✅ Usuário admin criado')
        else:
            self.stdout.write('✅ Usuário admin atualizado')
        
        # Criar ou atualizar profile como MASTER
        profile, created = UserProfile.objects.get_or_create(
            user=admin_user,
            defaults={
                'tipo_usuario': 'master',
                'is_master': True,
                'tipo_acesso': 'completo'
            }
        )
        
        # Sempre garantir que seja master
        profile.tipo_usuario = 'master'
        profile.is_master = True
        profile.tipo_acesso = 'completo'
        profile.save()
        
        if created:
            self.stdout.write('✅ Profile do admin criado como MASTER')
        else:
            self.stdout.write('✅ Profile do admin atualizado para MASTER')
    
    def importar_lojas(self):
        """Importa lojas do Excel"""
        # Verificar se já existem lojas
        if Loja.objects.exists():
            self.stdout.write('✅ Lojas já existem, pulando importação')
            return
            
        file_path = 'Lojas Portal.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write('⚠️ Arquivo Excel não encontrado, criando lojas básicas')
            self.criar_lojas_basicas()
            return
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            estado_map = {
                'SÃO PAULO': 'SP', 'RIO DE JANEIRO': 'RJ', 'MINAS GERAIS': 'MG',
                'BAHIA': 'BA', 'PARANÁ': 'PR', 'RIO GRANDE DO SUL': 'RS',
                'PERNAMBUCO': 'PE', 'CEARÁ': 'CE', 'PARÁ': 'PA',
                'SANTA CATARINA': 'SC', 'GOIÁS': 'GO', 'MARANHÃO': 'MA',
                'ESPÍRITO SANTO': 'ES', 'PIAUÍ': 'PI', 'ALAGOAS': 'AL',
                'TOCANTINS': 'TO', 'RIO GRANDE DO NORTE': 'RN', 'ACRE': 'AC',
                'AMAPÁ': 'AP', 'AMAZONAS': 'AM', 'MATO GROSSO': 'MT',
                'MATO GROSSO DO SUL': 'MS', 'RONDÔNIA': 'RO', 'RORAIMA': 'RR',
                'SERGIPE': 'SE', 'DISTRITO FEDERAL': 'DF', 'BLUMENAU': 'SC'
            }
            
            # Usar bulk_create para melhor performance
            lojas_para_criar = []
            importadas = 0
            
            for row in rows:
                if not row[0]:
                    continue
                
                try:
                    nome = f"Loja {row[0]}"
                    endereco = str(row[1]) if row[1] else ''
                    numero = str(row[2]) if row[2] else ''
                    municipio = str(row[4]) if row[4] else ''
                    estado = estado_map.get(str(row[5]).upper(), '')
                    cep = str(row[7]) if row[7] else ''
                    regional = str(row[8]) if row[8] else ''
                    
                    # Coordenadas
                    latitude = float(row[9]) if row[9] else None
                    longitude = float(row[10]) if row[10] else None
                    
                    if endereco and numero:
                        endereco = f"{endereco}, {numero}"
                    
                    lojas_para_criar.append(Loja(
                        nome=nome,
                        endereco=endereco,
                        numero='',
                        municipio=municipio,
                        estado=estado,
                        cep=cep,
                        regional=regional,
                        latitude=latitude,
                        longitude=longitude
                    ))
                    importadas += 1
                    
                except Exception as e:
                    continue
            
            # Criar todas as lojas de uma vez
            if lojas_para_criar:
                Loja.objects.bulk_create(lojas_para_criar, batch_size=100)
                self.stdout.write(f'✅ {importadas} lojas importadas')
            
        except Exception as e:
            self.stdout.write(f'⚠️ Erro ao importar lojas: {e}')
            self.criar_lojas_basicas()
    
    def criar_lojas_basicas(self):
        """Cria algumas lojas básicas se não conseguir importar do Excel"""
        lojas_basicas = [
            {
                'nome': 'Loja São Paulo',
                'endereco': 'Rua Augusta, 1000',
                'municipio': 'São Paulo',
                'estado': 'SP',
                'cep': '01305-100',
                'regional': 'SP'
            },
            {
                'nome': 'Loja Rio de Janeiro',
                'endereco': 'Av. Copacabana, 500',
                'municipio': 'Rio de Janeiro',
                'estado': 'RJ',
                'cep': '22020-000',
                'regional': 'RJ'
            },
            {
                'nome': 'Loja Belo Horizonte',
                'endereco': 'Rua da Bahia, 1200',
                'municipio': 'Belo Horizonte',
                'estado': 'MG',
                'cep': '30160-012',
                'regional': 'MG'
            }
        ]
        
        for loja_data in lojas_basicas:
            Loja.objects.get_or_create(
                nome=loja_data['nome'],
                defaults=loja_data
            )
        
        self.stdout.write('✅ Lojas básicas criadas')
    
    def cadastrar_transportadoras(self):
        """Cadastra transportadoras"""
        transportadoras_data = [
            {'nome': 'Log20', 'email': 'log20@lojasrenner.com.br'},
            {'nome': 'Soluciona', 'email': 'soluciona@lojasrenner.com.br'},
            {'nome': 'Leão Log', 'email': 'leaolog@lojasrenner.com.br'},
        ]
        
        cadastradas = 0
        for data in transportadoras_data:
            transportadora, created = Transportadora.objects.get_or_create(
                nome=data['nome'],
                defaults={'email': data['email']}
            )
            if created:
                cadastradas += 1
        
        self.stdout.write(f'✅ {cadastradas} transportadoras cadastradas')
