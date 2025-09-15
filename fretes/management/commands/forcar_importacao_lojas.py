from django.core.management.base import BaseCommand
from fretes.models import Loja
import os

class Command(BaseCommand):
    help = 'Força a importação das lojas do Excel'

    def handle(self, *args, **options):
        self.stdout.write('🔄 FORÇANDO IMPORTAÇÃO DAS LOJAS...')
        self.stdout.write('=' * 60)
        
        # Verificar se o arquivo existe
        file_path = 'Lojas Portal.xlsx'
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'❌ Arquivo {file_path} não encontrado!'))
            return
        
        try:
            # Usar openpyxl diretamente
            from openpyxl import load_workbook
            
            # Carregar o arquivo Excel
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            
            # Pular o cabeçalho (primeira linha)
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            
            self.stdout.write(f'📊 Arquivo Excel carregado: {len(rows)} linhas encontradas')
            
            # Limpar lojas existentes (opcional - comentado para não perder dados)
            # Loja.objects.all().delete()
            # self.stdout.write('🗑️ Lojas existentes removidas')
            
            # Contar lojas importadas
            importadas = 0
            existentes = 0
            erros = 0
            
            for i, row in enumerate(rows, start=2):
                if not row[0]:  # Se a primeira coluna estiver vazia, pular
                    continue
                    
                try:
                    # Extrair dados da linha - ajustar índices conforme estrutura real do Excel
                    nome = str(row[0])[:255] if row[0] else ''  # Limitar a 255 caracteres
                    endereco = str(row[1]) if row[1] else ''
                    numero = str(row[2])[:20] if row[2] else ''  # Limitar a 20 caracteres
                    municipio = str(row[3])[:100] if row[3] else ''  # Limitar a 100 caracteres
                    estado = str(row[4])[:10] if row[4] else ''  # Limitar a 10 caracteres
                    cep = str(row[5])[:15] if row[5] else ''  # Limitar a 15 caracteres
                    regional = str(row[6])[:100] if row[6] else ''  # Limitar a 100 caracteres
                    
                    # Tentar encontrar latitude e longitude - podem estar em colunas diferentes
                    latitude = None
                    longitude = None
                    
                    # Procurar por valores numéricos nas colunas restantes
                    for col_idx in range(7, len(row)):
                        if row[col_idx] is not None:
                            try:
                                valor = float(row[col_idx])
                                # Verificar se é uma coordenada válida
                                if -90 <= valor <= 90:  # Latitude válida
                                    if latitude is None:
                                        latitude = valor
                                elif -180 <= valor <= 180:  # Longitude válida
                                    if longitude is None:
                                        longitude = valor
                                elif latitude is None:  # Se não for coordenada válida, usar como latitude
                                    latitude = valor
                                elif longitude is None:  # Se não for coordenada válida, usar como longitude
                                    longitude = valor
                                    break
                            except (ValueError, TypeError):
                                continue
                    
                    if not nome:  # Pular se não tiver nome
                        continue
                    
                    # Criar ou atualizar loja
                    loja, created = Loja.objects.get_or_create(
                        nome=nome,
                        defaults={
                            'endereco': endereco,
                            'numero': numero,
                            'municipio': municipio,
                            'estado': estado,
                            'cep': cep,
                            'regional': regional,
                            'latitude': latitude,
                            'longitude': longitude
                        }
                    )
                    
                    if created:
                        importadas += 1
                        self.stdout.write(f'✅ Loja {loja.nome} importada')
                    else:
                        existentes += 1
                        # Atualizar dados se necessário
                        loja.endereco = endereco
                        loja.numero = numero
                        loja.municipio = municipio
                        loja.estado = estado
                        loja.cep = cep
                        loja.regional = regional
                        loja.latitude = latitude
                        loja.longitude = longitude
                        loja.save()
                        self.stdout.write(f'🔄 Loja {loja.nome} atualizada')
                        
                except Exception as e:
                    erros += 1
                    self.stdout.write(f'❌ Erro na linha {i}: {e}')
            
            self.stdout.write(f'\n📈 RESUMO DA IMPORTAÇÃO:')
            self.stdout.write(f'   - Lojas importadas: {importadas}')
            self.stdout.write(f'   - Lojas atualizadas: {existentes}')
            self.stdout.write(f'   - Erros: {erros}')
            self.stdout.write(f'   - Total no banco: {Loja.objects.count()}')
            
            # Verificar algumas lojas
            if Loja.objects.count() > 0:
                self.stdout.write(f'\n📋 PRIMEIRAS 5 LOJAS:')
                lojas = Loja.objects.all()[:5]
                for loja in lojas:
                    self.stdout.write(f'   - {loja.nome}: {loja.endereco}, {loja.municipio}/{loja.estado}')
                    if loja.latitude and loja.longitude:
                        self.stdout.write(f'     📍 Coordenadas: {loja.latitude}, {loja.longitude}')
            
            if importadas > 0 or existentes > 0:
                self.stdout.write(self.style.SUCCESS('🎉 Importação concluída com sucesso!'))
            else:
                self.stdout.write(self.style.WARNING('⚠️ Nenhuma loja foi processada'))
                
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'❌ Erro na importação: {e}'))
            import traceback
            self.stdout.write(traceback.format_exc())
        
        self.stdout.write('\n' + '=' * 60)
        self.stdout.write('🏁 IMPORTAÇÃO FORÇADA COMPLETA!')
